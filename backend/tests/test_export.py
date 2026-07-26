"""Tests for the Excel export endpoint."""
import io
import uuid
import json as _json
import re
import zipfile
import pytest
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

from routers.export import (
    _build_legacy_simple_report_pdf,
    _build_simple_report_pdf,
    _build_workbook,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def _seed_project_with_images(client):
    """Create a project with images, classifications, and comments for export testing."""
    resp = client.post("/api/projects/", json={
        "name": "Export Test Project",
        "description": "Project for testing Excel export",
        "meta_group_id": "test-group",
    })
    assert resp.status_code == 201, resp.text
    project = resp.json()

    class_resp = client.post(f"/api/projects/{project['id']}/classes", json={
        "name": "Defect",
        "description": "Visual defect detected",
        "project_id": project["id"],
    })
    assert class_resp.status_code == 201, class_resp.text
    defect_class = class_resp.json()

    class_resp2 = client.post(f"/api/projects/{project['id']}/classes", json={
        "name": "Scratch",
        "description": "Surface scratch",
        "project_id": project["id"],
    })
    assert class_resp2.status_code == 201, class_resp2.text
    scratch_class = class_resp2.json()

    image_ids = []
    for i in range(3):
        metadata = {
            "lot_number": f"LOT-{100 + i}",
            "part_serial_number": f"SN-{2000 + i}",
            "inspection_status": ["Not Reviewed", "Pass", "Reject"][i],
            "inspector_name": f"Inspector {i + 1}",
            "secondary_inspector_name": f"Secondary {i + 1}" if i > 0 else "",
        }
        files = {"file": (f"test_image_{i}.png", b"fake-png-data", "image/png")}
        data = {"metadata": _json.dumps(metadata)}
        img_resp = client.post(
            f"/api/projects/{project['id']}/images",
            files=files,
            data=data,
        )
        assert img_resp.status_code == 201, img_resp.text
        image_ids.append(img_resp.json()["id"])

    for img_id in image_ids[:2]:
        cl_resp = client.post(f"/api/images/{img_id}/classifications", json={
            "image_id": img_id,
            "class_id": defect_class["id"],
        })
        assert cl_resp.status_code == 201, cl_resp.text

    cl_resp2 = client.post(f"/api/images/{image_ids[1]}/classifications", json={
        "image_id": image_ids[1],
        "class_id": scratch_class["id"],
    })
    assert cl_resp2.status_code == 201, cl_resp2.text

    for idx, img_id in enumerate(image_ids):
        cmt_resp = client.post(f"/api/images/{img_id}/comments", json={
            "image_id": img_id,
            "text": f"Test comment for image {idx}",
        })
        assert cmt_resp.status_code == 201, cmt_resp.text

    return {
        "project": project,
        "image_ids": image_ids,
        "classes": [defect_class, scratch_class],
    }


# ---------------------------------------------------------------------------
# Endpoint integration tests
# ---------------------------------------------------------------------------

def test_export_excel_returns_xlsx(client, _seed_project_with_images):
    """Verify the export endpoint returns a valid Excel file."""
    project = _seed_project_with_images["project"]
    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    assert ".xlsx" in resp.headers["content-disposition"]

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.title == "Image Data"

    # Verify the total and check all expected headers are present
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers[0] == "Filename"
    assert "lot_number" in headers
    assert "part_serial_number" in headers
    assert "inspection_status" in headers
    assert "inspector_name" in headers
    assert "secondary_inspector_name" in headers
    assert "Review Status" in headers
    assert "Reviewer" in headers
    assert "Review Date" in headers
    assert "Image Classes" in headers
    assert "Comment" in headers

    assert ws.max_row == 4  # 1 header + 3 data rows


def test_export_excel_data_content(client, _seed_project_with_images):
    """Verify the actual data content in the exported Excel file."""
    project = _seed_project_with_images["project"]
    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    # Build a header -> column index map for robust lookups
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Row 2 = first image
    assert ws.cell(row=2, column=headers["Filename"]).value == "test_image_0.png"
    assert ws.cell(row=2, column=headers["lot_number"]).value == "LOT-100"
    assert ws.cell(row=2, column=headers["part_serial_number"]).value == "SN-2000"
    assert ws.cell(row=2, column=headers["inspection_status"]).value == "Not Reviewed"
    assert ws.cell(row=2, column=headers["inspector_name"]).value == "Inspector 1"

    # Row 3 = second image (has two classifications)
    assert ws.cell(row=3, column=headers["lot_number"]).value == "LOT-101"
    classes_val = ws.cell(row=3, column=headers["Image Classes"]).value
    assert "Defect" in classes_val
    assert "Scratch" in classes_val

    # Row 4 = third image (Reject status)
    assert ws.cell(row=4, column=headers["inspection_status"]).value == "Reject"


def test_export_excel_project_not_found(client):
    """Verify 404 for non-existent project."""
    fake_id = str(uuid.uuid4())
    resp = client.get(f"/api/projects/{fake_id}/export-excel")
    assert resp.status_code == 404


def test_export_excel_empty_project(client):
    """Verify export works for a project with no images."""
    resp = client.post("/api/projects/", json={
        "name": "Empty Project",
        "description": "No images",
        "meta_group_id": "test-group",
    })
    assert resp.status_code == 201
    project = resp.json()

    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 1


def test_export_excel_comments_concatenated(client, _seed_project_with_images):
    """Verify multiple comments on a single image are pipe-separated."""
    project = _seed_project_with_images["project"]
    img_id = _seed_project_with_images["image_ids"][0]

    # Add a second comment to the first image
    cmt_resp = client.post(f"/api/images/{img_id}/comments", json={
        "image_id": img_id,
        "text": "Second comment on image 0",
    })
    assert cmt_resp.status_code == 201

    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    comment_cell = ws.cell(row=2, column=headers["Comment"]).value
    assert "|" in comment_cell
    assert "Test comment for image 0" in comment_cell
    assert "Second comment on image 0" in comment_cell


def test_export_excel_deleted_images_excluded(client):
    """Verify that soft-deleted images are not included in the export."""
    resp = client.post("/api/projects/", json={
        "name": "Deletion Test",
        "description": "Test deleted images",
        "meta_group_id": "test-group",
    })
    project = resp.json()
    pid = project["id"]

    # Upload two images
    for i in range(2):
        files = {"file": (f"del_test_{i}.png", b"data", "image/png")}
        img_resp = client.post(f"/api/projects/{pid}/images", files=files)
        assert img_resp.status_code == 201

    # Soft-delete the first image via the correct endpoint
    images_resp = client.get(f"/api/projects/{pid}/images")
    images = images_resp.json()
    first_id = images[0]["id"]
    del_resp = client.request(
        "DELETE",
        f"/api/projects/{pid}/images/{first_id}",
        json={"reason": "test deletion"},
    )
    assert del_resp.status_code == 200, del_resp.text

    # Export should only contain the non-deleted image
    resp = client.get(f"/api/projects/{pid}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active
    assert ws.max_row == 2  # 1 header + 1 data row


def test_export_excel_no_metadata_defaults(client):
    """Verify images with no metadata export only fixed columns."""
    resp = client.post("/api/projects/", json={
        "name": "No Metadata",
        "description": "Images without metadata",
        "meta_group_id": "test-group",
    })
    project = resp.json()

    files = {"file": ("bare_image.png", b"data", "image/png")}
    client.post(f"/api/projects/{project['id']}/images", files=files)

    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active

    # With no metadata, 6 columns: Filename, Review Status, Reviewer, Review Date,
    # Image Classes, Comment
    assert ws.max_column == 6
    headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    assert headers[0] == "Filename"
    assert "Review Status" in headers
    assert "Reviewer" in headers
    assert "Review Date" in headers
    assert "Image Classes" in headers
    assert "Comment" in headers

    # Filename should be present
    assert ws.cell(row=2, column=1).value == "bare_image.png"


def test_export_excel_filename_sanitization(client):
    """Verify special characters in project name are sanitized in the filename."""
    resp = client.post("/api/projects/", json={
        "name": "Test/Project <with> special&chars!",
        "description": "Special chars test",
        "meta_group_id": "test-group",
    })
    project = resp.json()

    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    disposition = resp.headers["content-disposition"]
    # Should not contain path separators or angle brackets
    assert "/" not in disposition.split("filename=")[1]
    assert "<" not in disposition.split("filename=")[1]
    assert ">" not in disposition.split("filename=")[1]
    assert ".xlsx" in disposition


def test_export_excel_image_no_classifications(client):
    """Verify images with no classifications have empty Image Classes column."""
    resp = client.post("/api/projects/", json={
        "name": "No Classes Assigned",
        "description": "Test",
        "meta_group_id": "test-group",
    })
    project = resp.json()

    files = {"file": ("unclassified.png", b"data", "image/png")}
    client.post(f"/api/projects/{project['id']}/images", files=files)

    resp = client.get(f"/api/projects/{project['id']}/export-excel")

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    classes_val = ws.cell(row=2, column=headers["Image Classes"]).value
    assert classes_val in (None, "")


def test_export_excel_image_no_comments(client):
    """Verify images with no comments have empty Comment column."""
    resp = client.post("/api/projects/", json={
        "name": "No Comments",
        "description": "Test",
        "meta_group_id": "test-group",
    })
    project = resp.json()

    files = {"file": ("silent.png", b"data", "image/png")}
    client.post(f"/api/projects/{project['id']}/images", files=files)

    resp = client.get(f"/api/projects/{project['id']}/export-excel")

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    comment_val = ws.cell(row=2, column=headers["Comment"]).value
    assert comment_val in (None, "")


def test_export_excel_forbidden_for_non_group_member(client):
    """Verify 403 when user is not a member of the project's group."""
    resp = client.post("/api/projects/", json={
        "name": "Restricted Project",
        "description": "Not for everyone",
        "meta_group_id": "restricted-group",
    })
    assert resp.status_code == 201
    project = resp.json()

    with patch("routers.export.is_user_in_group", return_value=False):
        resp = client.get(f"/api/projects/{project['id']}/export-excel")

    assert resp.status_code == 403
    assert "access" in resp.json()["detail"].lower()


def test_export_excel_alternate_metadata_keys(client):
    """Verify that all metadata keys become their own columns in the export."""
    resp = client.post("/api/projects/", json={
        "name": "Alternate Keys",
        "description": "Test alternate metadata key names",
        "meta_group_id": "test-group",
    })
    project = resp.json()

    # Use arbitrary key names - each should become its own column
    metadata = {
        "lotNumber": "ALT-LOT-1",
        "serial": "ALT-SN-1",
        "inspectionStatus": "Pass",
        "inspectorName": "Alt Inspector",
        "secondaryInspectorName": "Alt Secondary",
    }
    files = {"file": ("alt_keys.png", b"data", "image/png")}
    data = {"metadata": _json.dumps(metadata)}
    client.post(f"/api/projects/{project['id']}/images", files=files, data=data)

    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active

    # Build a header -> column map
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Each metadata key should appear as its own column header
    assert "lotNumber" in headers
    assert "serial" in headers
    assert "inspectionStatus" in headers
    assert "inspectorName" in headers
    assert "secondaryInspectorName" in headers

    # Values should be present in the correct columns
    assert ws.cell(row=2, column=headers["lotNumber"]).value == "ALT-LOT-1"
    assert ws.cell(row=2, column=headers["serial"]).value == "ALT-SN-1"
    assert ws.cell(row=2, column=headers["inspectionStatus"]).value == "Pass"
    assert ws.cell(row=2, column=headers["inspectorName"]).value == "Alt Inspector"
    assert ws.cell(row=2, column=headers["secondaryInspectorName"]).value == "Alt Secondary"


def test_export_excel_includes_review_data(client):
    """Verify review status, reviewer, and review date appear in the export."""
    # Create a project and image
    proj_resp = client.post("/api/projects/", json={
        "name": "Review Data Test",
        "description": "Test review columns in export",
        "meta_group_id": "test-group",
    })
    assert proj_resp.status_code == 201
    project = proj_resp.json()

    img_resp = client.post(
        f"/api/projects/{project['id']}/images",
        files={"file": ("reviewed.png", b"data", "image/png")},
    )
    assert img_resp.status_code == 201
    image_id = img_resp.json()["id"]

    # Create a review for the image
    rev_resp = client.post(f"/api/images/{image_id}/reviews", json={"status": "pass"})
    assert rev_resp.status_code == 201

    # Export and check
    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    assert "Review Status" in headers
    assert "Reviewer" in headers
    assert "Review Date" in headers

    assert ws.cell(row=2, column=headers["Review Status"]).value == "pass"
    reviewer_val = ws.cell(row=2, column=headers["Reviewer"]).value
    assert reviewer_val is not None and reviewer_val != ""
    review_date_val = ws.cell(row=2, column=headers["Review Date"]).value
    assert review_date_val is not None and review_date_val != ""


def test_export_excel_no_review_shows_empty(client):
    """Verify images with no review have empty review columns."""
    proj_resp = client.post("/api/projects/", json={
        "name": "No Review Test",
        "description": "Test empty review columns",
        "meta_group_id": "test-group",
    })
    assert proj_resp.status_code == 201
    project = proj_resp.json()

    client.post(
        f"/api/projects/{project['id']}/images",
        files={"file": ("unreviewed.png", b"data", "image/png")},
    )

    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    assert ws.cell(row=2, column=headers["Review Status"]).value in (None, "")
    assert ws.cell(row=2, column=headers["Reviewer"]).value in (None, "")
    assert ws.cell(row=2, column=headers["Review Date"]).value in (None, "")


def test_export_excel_excludes_measurements_metadata(client):
    """Verify the 'measurements' metadata key is not exported as a column."""
    import json as _json

    proj_resp = client.post("/api/projects/", json={
        "name": "Measurements Excluded",
        "description": "Test that measurements key is excluded",
        "meta_group_id": "test-group",
    })
    assert proj_resp.status_code == 201
    project = proj_resp.json()

    metadata = {
        "lot_number": "LOT-1",
        "measurements": [{"id": "m1", "distance_pixels": 42.5}],
    }
    files = {"file": ("img.png", b"data", "image/png")}
    data = {"metadata": _json.dumps(metadata)}
    img_resp = client.post(
        f"/api/projects/{project['id']}/images",
        files=files,
        data=data,
    )
    assert img_resp.status_code == 201

    resp = client.get(f"/api/projects/{project['id']}/export-excel")
    assert resp.status_code == 200

    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(resp.content)).active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    assert "measurements" not in headers
    assert "lot_number" in headers


@pytest.mark.smoke
def test_project_json_report_v3_has_one_ordered_row_per_part_and_explicit_v2_compatibility(client):
    headers = {
        "X-User-Id": "report-v2@example.com",
        "X-User-Groups": "[\"report-v2-group\"]",
    }
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": "Concise Inspection Report",
            "description": "v3 report contract",
            "meta_group_id": "report-v2-group",
            "project_type": "PT2",
        },
        headers=headers,
    )
    assert project_resp.status_code == 201, project_resp.text
    project = project_resp.json()

    raw_parts = (
        ("SERIAL-Z", "unreviewed", "unreviewed", False),
        ("SERIAL-A", "in_review", "unreviewed", False),
        ("SERIAL-M", "pass", "pass", True),
        ("SERIAL-B", "reject_pending", "reject", True),
        ("SERIAL-C", "reject_confirmed", "reject", True),
    )
    created_parts = {}
    for serial_number, review_state, _inspection_result, _reviewed in raw_parts:
        part_resp = client.post(
            f"/api/projects/{project['id']}/parts",
            json={
                "serial_number": serial_number,
                "display_name": serial_number,
                "review_state": review_state,
            },
            headers=headers,
        )
        assert part_resp.status_code == 201, part_resp.text
        created_parts[serial_number] = part_resp.json()

    # Multiple images for one part must not duplicate that part's report row.
    pass_part_id = created_parts["SERIAL-M"]["id"]
    for index in range(2):
        image_resp = client.post(
            f"/api/projects/{project['id']}/images",
            files={"file": (f"duplicate-source-{index}.png", b"image-bytes", "image/png")},
            data={"metadata": _json.dumps({"part_id": pass_part_id})},
            headers=headers,
        )
        assert image_resp.status_code == 201, image_resp.text

    report_resp = client.get(f"/api/projects/{project['id']}/report-json", headers=headers)
    assert report_resp.status_code == 200, report_resp.text
    payload = report_resp.json()

    assert set(payload) == {"schema_version", "project", "summary", "parts"}
    assert payload["schema_version"] == 3
    assert set(payload["project"]) == {"id", "name", "project_type", "meta_group_id"}
    assert payload["project"] == {
        "id": project["id"],
        "name": "Concise Inspection Report",
        "project_type": "PT2",
        "meta_group_id": "report-v2-group",
    }
    assert set(payload["summary"]) == {
        "total_parts",
        "reviewed_parts",
        "unreviewed_parts",
        "part_status_counts",
    }
    assert payload["summary"] == {
        "total_parts": 5,
        "reviewed_parts": 3,
        "unreviewed_parts": 2,
        "part_status_counts": {"pass": 1, "reject": 2, "unreviewed": 2},
    }

    expected_by_serial = {
        serial_number: inspection_result
        for serial_number, _review_state, inspection_result, _reviewed in raw_parts
    }
    assert [row["part_identifier"] for row in payload["parts"]] == [
        "SERIAL-A",
        "SERIAL-B",
        "SERIAL-C",
        "SERIAL-M",
        "SERIAL-Z",
    ]
    assert len(payload["parts"]) == len(created_parts)
    assert len({row["part_id"] for row in payload["parts"]}) == len(created_parts)
    for row in payload["parts"]:
        assert set(row) == {"part_id", "part_identifier", "inspection_result"}
        assert row["part_id"] == created_parts[row["part_identifier"]]["id"]
        assert row["inspection_result"] == expected_by_serial[row["part_identifier"]]

    v2_resp = client.get(
        f"/api/projects/{project['id']}/report-json?schema_version=2",
        headers=headers,
    )
    assert v2_resp.status_code == 200, v2_resp.text
    v2_payload = v2_resp.json()
    assert set(v2_payload) == {
        "project",
        "summary",
        "part_assignments",
        "part_review_summary",
        "image_part_mappings",
    }
    assert v2_payload["summary"]["total_images"] == 2
    assert v2_payload["summary"]["total_batches"] == 0
    assert v2_payload["summary"]["total_parts"] == 5
    assert v2_payload["summary"]["reviewed_parts"] == 3
    assert v2_payload["summary"]["metadata_normalization"] == {
        "dropped_non_object_items": {
            "annotations": 0,
            "overlay_layers": 0,
            "segmentation_runs": 0,
            "measurement_runs": 0,
        }
    }
    assert len(v2_payload["part_assignments"]) == 5
    assert len(v2_payload["part_review_summary"]) == 5
    assert len(v2_payload["image_part_mappings"]) == 2
    assert {
        row["filename"] for row in v2_payload["image_part_mappings"]
    } == {"duplicate-source-0.png", "duplicate-source-1.png"}


def test_project_json_report_v3_empty_project(client):
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": "Empty Inspection Report",
            "description": "no parts",
            "meta_group_id": "test-group",
            "project_type": "PT3",
        },
    )
    assert project_resp.status_code == 201, project_resp.text
    project = project_resp.json()

    report_resp = client.get(f"/api/projects/{project['id']}/report-json")
    assert report_resp.status_code == 200, report_resp.text
    payload = report_resp.json()

    assert payload == {
        "schema_version": 3,
        "project": {
            "id": project["id"],
            "name": "Empty Inspection Report",
            "project_type": "PT3",
            "meta_group_id": "test-group",
        },
        "summary": {
            "total_parts": 0,
            "reviewed_parts": 0,
            "unreviewed_parts": 0,
            "part_status_counts": {"pass": 0, "reject": 0, "unreviewed": 0},
        },
        "parts": [],
    }


def test_project_json_report_forbidden_for_non_group_member(client):
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": "JSON Report Access",
            "description": "access check",
            "meta_group_id": "json-report-private",
            "project_type": "PT1",
        },
    )
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    with patch("routers.export.is_user_in_group", return_value=False):
        report_resp = client.get(f"/api/projects/{project_id}/report-json")

    assert report_resp.status_code == 403


def _make_report_payload(part_rows, schema_version=3):
    parts = [
        {
            "part_id": str(uuid.uuid4()),
            "part_identifier": identifier,
            "inspection_result": inspection_result,
        }
        for identifier, inspection_result in part_rows
    ]
    if schema_version == 2:
        for part in parts:
            part["reviewed"] = part["inspection_result"] != "unreviewed"
    status_counts = {
        status: sum(part["inspection_result"] == status for part in parts)
        for status in ("pass", "reject", "unreviewed")
    }
    return {
        "schema_version": schema_version,
        "project": {
            "id": str(uuid.uuid4()),
            "name": "PDF Contract Project",
            "project_type": "PT1",
            "meta_group_id": "pdf-report-group",
        },
        "summary": {
            "total_parts": len(parts),
            "reviewed_parts": status_counts["pass"] + status_counts["reject"],
            "unreviewed_parts": status_counts["unreviewed"],
            "part_status_counts": status_counts,
        },
        "parts": parts,
    }


def _pdf_page_count(pdf_bytes):
    return len(re.findall(rb"/Type /Page\b", pdf_bytes))


def _pdf_content_streams(pdf_bytes):
    return re.findall(rb"stream\n(.*?)\nendstream", pdf_bytes, flags=re.DOTALL)


def _pdf_extracted_lines(pdf_bytes):
    lines = []
    for stream in _pdf_content_streams(pdf_bytes):
        for encoded_line in re.findall(rb"^\((.*)\) Tj$", stream, flags=re.MULTILINE):
            line = (
                encoded_line
                .replace(rb"\\(", b"(")
                .replace(rb"\\)", b")")
                .replace(b"\\\\", b"\\")
            )
            lines.append(line.decode("latin-1"))
    return lines


def test_simple_pdf_report_renders_canonical_mixed_part_rows_once():
    pdf_bytes = _build_simple_report_pdf(
        _make_report_payload(
            [
                ("PART-PASS", "pass"),
                ("PART-REJECT", "reject"),
                ("PART-UNREVIEWED", "unreviewed"),
            ]
        )
    )

    assert _pdf_page_count(pdf_bytes) == 1
    assert b"Part" in pdf_bytes
    assert b"Result" in pdf_bytes
    table_headers = [line for line in _pdf_extracted_lines(pdf_bytes) if line.startswith("Part ")]
    assert len(table_headers) == 1
    assert "Reviewed" not in table_headers[0]
    assert pdf_bytes.count(b"PART-PASS") == 1
    assert pdf_bytes.count(b"PART-REJECT") == 1
    assert pdf_bytes.count(b"PART-UNREVIEWED") == 1
    assert b"Part Pass/Fail Assignments" not in pdf_bytes
    assert b"Part Status Summary" not in pdf_bytes
    assert b"Image-to-Part Mapping" not in pdf_bytes


def test_legacy_pdf_report_preserves_assignments_summary_and_mapping_pages():
    payload = {
        "project": {"id": "project-1", "name": "Legacy", "project_type": "PT1"},
        "summary": {
            "total_images": 1,
            "total_batches": 1,
            "total_parts": 1,
            "reviewed_parts": 1,
            "unreviewed_parts": 0,
            "part_status_counts": {"pass": 1, "reject": 0, "unreviewed": 0},
            "metadata_normalization": {"dropped_non_object_items": {}},
        },
        "part_assignments": [{
            "part_identifier": "PART-PASS",
            "pass_fail": "pass",
            "username": "inspector",
            "batch_owner": "owner",
            "assigned_at": "2026-01-01",
        }],
        "part_review_summary": [{
            "part_identifier": "PART-PASS",
            "review_status": "pass",
        }],
        "image_part_mappings": [{
            "filename": "source.png",
            "part_identifier": "PART-PASS",
        }],
    }
    pdf_bytes = _build_legacy_simple_report_pdf(payload)

    assert _pdf_page_count(pdf_bytes) == 2
    assert b"Part Pass/Fail Assignments" in pdf_bytes
    assert b"Part Status Summary" in pdf_bytes
    assert b"VISTA Report Image-to-Part Mapping" in pdf_bytes
    assert b"source.png -> PART-PASS" in pdf_bytes


def test_simple_pdf_report_renders_empty_part_table():
    pdf_bytes = _build_simple_report_pdf(_make_report_payload([]))

    assert _pdf_page_count(pdf_bytes) == 1
    assert b"Total Parts: 0" in pdf_bytes
    assert b"\\(no parts\\)" in pdf_bytes
    assert b"Page 1 of 1" in pdf_bytes


def test_simple_pdf_report_paginates_every_part_without_truncation():
    part_rows = [
        (f"PAGE-PART-{index:03d}", ("pass", "reject", "unreviewed")[index % 3])
        for index in range(85)
    ]
    pdf_bytes = _build_simple_report_pdf(_make_report_payload(part_rows))

    assert _pdf_page_count(pdf_bytes) == 3
    assert b"/Count 3" in pdf_bytes
    assert pdf_bytes.count(b"PAGE-PART-000") == 1
    assert pdf_bytes.count(b"PAGE-PART-042") == 1
    assert pdf_bytes.count(b"PAGE-PART-084") == 1
    assert sum(pdf_bytes.count(f"PAGE-PART-{index:03d}".encode()) for index in range(85)) == 85
    assert b"Page 3 of 3" in pdf_bytes


def test_simple_pdf_report_escapes_and_wraps_special_long_identifier():
    special_identifier = "PART (A) \\ control\nline\x01 café " + ("X" * 150)
    pdf_bytes = _build_simple_report_pdf(
        _make_report_payload([(special_identifier, "reject")])
    )

    assert _pdf_page_count(pdf_bytes) == 1
    assert b"PART \\(A\\) \\\\ control line  caf\xe9" in pdf_bytes
    assert pdf_bytes.count(b"X") == 150
    assert pdf_bytes.endswith(b"%%EOF")


def test_simple_pdf_report_preserves_distinct_non_latin_identifiers_as_ascii_escapes():
    pdf_bytes = _build_simple_report_pdf(
        _make_report_payload(
            [
                ("部品一", "pass"),
                ("部品二", "reject"),
            ]
        )
    )
    extracted_text = "\n".join(_pdf_extracted_lines(pdf_bytes))

    assert r"\u90e8\u54c1\u4e00" in extracted_text
    assert r"\u90e8\u54c1\u4e8c" in extracted_text
    assert r"\u90e8\u54c1\u4e00" != r"\u90e8\u54c1\u4e8c"
    assert b"?" not in pdf_bytes


def test_simple_pdf_report_wraps_max_length_project_header():
    payload = _make_report_payload([("PART-001", "pass")])
    payload["project"]["name"] = "P" * 255

    extracted_lines = _pdf_extracted_lines(_build_simple_report_pdf(payload))
    project_lines = [
        line.removeprefix("Project: ")
        for line in extracted_lines
        if line.startswith("Project: ") or (line and set(line) == {"P"})
    ]

    assert "".join(project_lines) == "P" * 255
    assert all(len(line) <= 84 for line in extracted_lines)


def test_simple_pdf_report_moves_a_wrapped_logical_row_intact_to_the_next_page():
    short_rows = [(f"SHORT-{index:02d}", "pass") for index in range(39)]
    long_identifier = "BOUNDARY-" + ("X" * 80)
    pdf_bytes = _build_simple_report_pdf(
        _make_report_payload([*short_rows, (long_identifier, "reject")])
    )
    content_streams = _pdf_content_streams(pdf_bytes)

    assert _pdf_page_count(pdf_bytes) == 2
    assert b"BOUNDARY-" not in content_streams[0]
    assert b"BOUNDARY-" in content_streams[1]
    assert content_streams[1].count(b"reject") == 1
    assert b"yes" not in content_streams[1]


def test_simple_pdf_report_repeats_disposition_when_one_row_spans_pages():
    huge_identifier = "HUGE-" + ("Z" * 5000)
    content_streams = _pdf_content_streams(
        _build_simple_report_pdf(_make_report_payload([(huge_identifier, "reject")]))
    )

    assert len(content_streams) >= 2
    assert all(b"reject" in stream and b"yes" not in stream for stream in content_streams)
    assert sum(stream.count(b"Z") for stream in content_streams) == 5000


def test_project_pdf_report_supports_export(client):
    headers = {
        "X-User-Id": "pdf-report@example.com",
        "X-User-Groups": "[\"pdf-report-group\"]",
    }
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": "PDF Report Project",
            "description": "pdf report scenario",
            "meta_group_id": "pdf-report-group",
            "project_type": "PT1",
        },
        headers=headers,
    )
    assert project_resp.status_code == 201, project_resp.text
    project_id = project_resp.json()["id"]

    report_resp = client.get(f"/api/projects/{project_id}/report-pdf", headers=headers)
    assert report_resp.status_code == 200, report_resp.text
    assert report_resp.headers["content-type"].startswith("application/pdf")
    assert 'filename="PDF Report Project-report.pdf"' in report_resp.headers["content-disposition"]
    assert "filename*=UTF-8''PDF%20Report%20Project-report.pdf" in report_resp.headers["content-disposition"]
    assert report_resp.content.startswith(b"%PDF-1.4")
    v3_headers = [
        line for line in _pdf_extracted_lines(report_resp.content) if line.startswith("Part ")
    ]
    assert len(v3_headers) == 1
    assert "Reviewed" not in v3_headers[0]

    v2_report_resp = client.get(
        f"/api/projects/{project_id}/report-pdf?schema_version=2",
        headers=headers,
    )
    assert v2_report_resp.status_code == 200, v2_report_resp.text
    assert _pdf_page_count(v2_report_resp.content) == 2
    assert b"Part Pass/Fail Assignments" in v2_report_resp.content
    assert b"VISTA Report Image-to-Part Mapping" in v2_report_resp.content

    for endpoint in ("report-json", "report-pdf"):
        unsupported_resp = client.get(
            f"/api/projects/{project_id}/{endpoint}?schema_version=4",
            headers=headers,
        )
        assert unsupported_resp.status_code == 422
        assert unsupported_resp.json()["detail"] == "schema_version must be 2 or 3"


def test_project_pdf_report_uses_safe_unicode_content_disposition(client):
    project_name = 'Résumé "部品"/QA\r\nX-Injected: yes'
    headers = {
        "X-User-Id": "pdf-header@example.com",
        "X-User-Groups": "[\"pdf-header-group\"]",
    }
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": project_name,
            "description": "pdf header hardening",
            "meta_group_id": "pdf-header-group",
            "project_type": "PT1",
        },
        headers=headers,
    )
    assert project_resp.status_code == 201, project_resp.text

    report_resp = client.get(
        f"/api/projects/{project_resp.json()['id']}/report-pdf",
        headers=headers,
    )

    assert report_resp.status_code == 200, report_resp.text
    disposition = report_resp.headers["content-disposition"]
    assert disposition.startswith('attachment; filename="')
    assert "\r" not in disposition and "\n" not in disposition
    assert "/" not in disposition and '\\"' not in disposition
    assert "filename*=UTF-8''" in disposition
    assert "%C3%A9" in disposition
    assert "%E9%83%A8%E5%93%81" in disposition
    assert "%0D" not in disposition and "%0A" not in disposition


def test_project_report_with_images_pdf_endpoint_auth_filename_and_content(client):
    project_name = 'Evidence Résumé "部品"/QA\r\nInjected'
    headers = {
        "X-User-Id": "evidence-report@example.com",
        "X-User-Groups": "[\"evidence-report-group\"]",
    }
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": project_name,
            "description": "image evidence report",
            "meta_group_id": "evidence-report-group",
            "project_type": "PT1",
        },
        headers=headers,
    )
    assert project_resp.status_code == 201, project_resp.text
    project_id = project_resp.json()["id"]
    synthetic_pdf = b"%PDF-1.4\nsynthetic evidence\n%%EOF"

    with patch(
        "routers.export.build_project_report_with_images_pdf",
        new=AsyncMock(
            return_value=SimpleNamespace(
                pdf_bytes=synthetic_pdf,
                page_count=1,
                panel_count=0,
                omissions=[],
            )
        ),
    ):
        report_resp = client.get(
            f"/api/projects/{project_id}/report-with-images-pdf",
            headers=headers,
        )

    assert report_resp.status_code == 200, report_resp.text
    assert report_resp.content == synthetic_pdf
    assert report_resp.headers["content-type"].startswith("application/pdf")
    disposition = report_resp.headers["content-disposition"]
    assert "-report-with-images.pdf" in disposition
    assert "filename*=UTF-8''" in disposition
    assert "%C3%A9" in disposition
    assert "%E9%83%A8%E5%93%81" in disposition
    assert "\r" not in disposition and "\n" not in disposition

    with patch("routers.export.is_user_in_group", return_value=False):
        forbidden_resp = client.get(
            f"/api/projects/{project_id}/report-with-images-pdf",
            headers={"X-User-Id": "outsider@example.com", "X-User-Groups": "[]"},
        )
    assert forbidden_resp.status_code == 403


def test_project_bundle_json_supports_progressive_users_per_project_type(client):
    import json as _json

    project_types = ("PT1", "PT2", "PT3")
    scenarios = (
        {
            "user": "basic",
            "level": 1,
            "part_count": 1,
            "image_count": 1,
            "overlay_layers": ["mask_base"],
            "annotation_count": 1,
            "segmentation_runs": 1,
            "measurement_runs": 1,
        },
        {
            "user": "intermediate",
            "level": 2,
            "part_count": 2,
            "image_count": 2,
            "overlay_layers": ["mask_base", "heatmap"],
            "annotation_count": 2,
            "segmentation_runs": 2,
            "measurement_runs": 2,
        },
        {
            "user": "advanced",
            "level": 3,
            "part_count": 3,
            "image_count": 3,
            "overlay_layers": ["mask_base", "heatmap", "depth"],
            "annotation_count": 3,
            "segmentation_runs": 3,
            "measurement_runs": 3,
        },
    )

    for project_type in project_types:
        for scenario in scenarios:
            group = f"bundle-json-{project_type}-{scenario['user']}"
            headers = {"X-Forwarded-Email": f"{scenario['user']}@{group}.test"}
            project_resp = client.post(
                "/api/projects/",
                json={
                    "name": f"Bundle JSON {project_type} {scenario['user']}",
                    "description": "bundle export coverage",
                    "meta_group_id": group,
                    "project_type": project_type,
                },
                headers=headers,
            )
            assert project_resp.status_code == 201, project_resp.text
            project_id = project_resp.json()["id"]

            total_annotations = 0
            total_overlay_layers = 0
            total_segmentation_runs = 0
            total_measurement_runs = 0

            for idx in range(scenario["part_count"]):
                batch_resp = client.post(
                    f"/api/projects/{project_id}/batches",
                    json={"name": f"batch-{idx}", "description": f"batch {idx}"},
                    headers=headers,
                )
                assert batch_resp.status_code == 201, batch_resp.text
                batch_id = batch_resp.json()["id"]

                annotations = [
                    {
                        "id": f"ann-{idx}-{annotation_idx}",
                        "defect_class": "scratch",
                        "modality": "rgb",
                        "comment": f"annotation-{annotation_idx}",
                    }
                    for annotation_idx in range(scenario["annotation_count"])
                ]
                if scenario["level"] > 1 and idx == 0:
                    annotations[0]["modality"] = ""
                segmentation_runs = [
                    {"overlay_id": f"overlay-{idx}-{run_idx}"}
                    for run_idx in range(scenario["segmentation_runs"])
                ]
                measurement_runs = [
                    {"run_id": f"measure-{idx}-{run_idx}"}
                    for run_idx in range(scenario["measurement_runs"])
                ]
                part_metadata = {
                    "overlay_layers": scenario["overlay_layers"],
                    "annotations": annotations,
                    "segmentation_runs": segmentation_runs,
                    "measurement_runs": measurement_runs,
                }
                total_annotations += len(annotations)
                total_overlay_layers += len(scenario["overlay_layers"])
                total_segmentation_runs += len(segmentation_runs)
                total_measurement_runs += len(measurement_runs)

                part_resp = client.post(
                    f"/api/projects/{project_id}/parts",
                    json={
                        "serial_number": f"{project_type}-{scenario['level']}-{idx}",
                        "display_name": f"part-{idx}",
                        "batch_id": batch_id,
                        "metadata": part_metadata,
                    },
                    headers=headers,
                )
                assert part_resp.status_code == 201, part_resp.text

            for image_idx in range(scenario["image_count"]):
                files = {
                    "file": (
                        f"{project_type}_{scenario['user']}_{image_idx}.png",
                        b"synthetic-image-data",
                        "image/png",
                    )
                }
                image_resp = client.post(
                    f"/api/projects/{project_id}/images",
                    files=files,
                    data={"metadata": _json.dumps({"slot": image_idx, "scenario": scenario["user"]})},
                    headers=headers,
                )
                assert image_resp.status_code == 201, image_resp.text

            bundle_resp = client.get(f"/api/projects/{project_id}/export-bundle-json", headers=headers)
            assert bundle_resp.status_code == 200, bundle_resp.text
            payload = bundle_resp.json()
            assert payload["project"]["project_type"] == project_type
            assert payload["bundle_summary"]["images"]["total"] == scenario["image_count"]
            assert payload["bundle_summary"]["parts"]["total"] == scenario["part_count"]
            assert payload["bundle_summary"]["annotations"]["total"] == total_annotations
            assert len(payload["bundle_summary"]["annotations"]["records"]) == total_annotations
            assert payload["bundle_summary"]["overlays"]["configured_layers"] == total_overlay_layers
            assert payload["bundle_summary"]["overlays"]["segmentation_runs"] == total_segmentation_runs
            assert len(payload["bundle_summary"]["overlays"]["records"]) == total_overlay_layers
            assert payload["bundle_summary"]["measurements"]["ai_runs"] == total_measurement_runs
            assert len(payload["bundle_summary"]["measurements"]["records"]) == total_measurement_runs
            assert payload["bundle_summary"]["images"]["total_bytes"] > 0
            assert len(payload["bundle_summary"]["discrepancies"]["per_part"]) == scenario["part_count"]
            expected_discrepancy_parts = 1 if scenario["level"] > 1 else 0
            assert payload["bundle_summary"]["discrepancies"]["parts_with_discrepancies"] == expected_discrepancy_parts


@pytest.mark.parametrize("project_type", ["PT1", "PT2", "PT3"])
def test_project_bundle_json_ignores_adversarial_non_list_metadata_shapes(client, project_type):
    group = f"bundle-json-adversarial-{project_type.lower()}"
    headers = {"X-Forwarded-Email": f"adversary@{group}.test"}

    project_resp = client.post(
        "/api/projects/",
        json={
            "name": f"Bundle JSON adversarial {project_type}",
            "description": "adversarial metadata shape test",
            "meta_group_id": group,
            "project_type": project_type,
        },
        headers=headers,
    )
    assert project_resp.status_code == 201, project_resp.text
    project_id = project_resp.json()["id"]

    batch_resp = client.post(
        f"/api/projects/{project_id}/batches",
        json={"name": "batch-0", "description": "adversarial batch"},
        headers=headers,
    )
    assert batch_resp.status_code == 201, batch_resp.text
    batch_id = batch_resp.json()["id"]

    part_resp = client.post(
        f"/api/projects/{project_id}/parts",
        json={
            "serial_number": f"{project_type}-ADV-0001",
            "display_name": "adversarial-part",
            "batch_id": batch_id,
            "metadata": {
                "annotations": "not-a-list",
                "overlay_layers": {"id": "overlay-1"},
                "segmentation_runs": 404,
                "measurement_runs": True,
            },
        },
        headers=headers,
    )
    assert part_resp.status_code == 201, part_resp.text

    bundle_resp = client.get(f"/api/projects/{project_id}/export-bundle-json", headers=headers)
    assert bundle_resp.status_code == 200, bundle_resp.text
    payload = bundle_resp.json()

    assert payload["project"]["project_type"] == project_type
    assert payload["bundle_summary"]["parts"]["total"] == 1
    assert payload["bundle_summary"]["annotations"]["total"] == 0
    assert payload["bundle_summary"]["overlays"]["configured_layers"] == 0
    assert payload["bundle_summary"]["overlays"]["segmentation_runs"] == 0
    assert payload["bundle_summary"]["measurements"]["ai_runs"] == 0
    assert payload["bundle_summary"]["annotations"]["records"] == []
    assert payload["bundle_summary"]["overlays"]["records"] == []
    assert payload["bundle_summary"]["measurements"]["records"] == []
    assert payload["bundle_summary"]["discrepancies"]["parts_with_discrepancies"] == 0


@pytest.mark.parametrize("project_type", ["PT1", "PT2", "PT3"])
def test_project_bundle_json_flags_dropped_non_object_metadata_items(client, project_type):
    group = f"bundle-json-dropped-{project_type.lower()}"
    headers = {"X-Forwarded-Email": f"normalizer@{group}.test"}

    project_resp = client.post(
        "/api/projects/",
        json={
            "name": f"Bundle JSON dropped items {project_type}",
            "description": "mixed list metadata hardening",
            "meta_group_id": group,
            "project_type": project_type,
        },
        headers=headers,
    )
    assert project_resp.status_code == 201, project_resp.text
    project_id = project_resp.json()["id"]

    batch_resp = client.post(
        f"/api/projects/{project_id}/batches",
        json={"name": "batch-0", "description": "mixed-metadata batch"},
        headers=headers,
    )
    assert batch_resp.status_code == 201, batch_resp.text
    batch_id = batch_resp.json()["id"]

    part_resp = client.post(
        f"/api/projects/{project_id}/parts",
        json={
            "serial_number": f"{project_type}-MIXED-0001",
            "display_name": "mixed-list-part",
            "batch_id": batch_id,
            "metadata": {
                "annotations": [{"id": "ann-1", "defect_class": "scratch", "modality": "rgb"}, "junk", 5],
                "overlay_layers": [{"id": "overlay-1", "label": "Mask", "color": "#22c55e"}, 9],
                "segmentation_runs": [{"overlay_id": "overlay-1"}, None],
                "measurement_runs": [{"run_id": "run-1", "status": "completed"}, False],
            },
        },
        headers=headers,
    )
    assert part_resp.status_code == 201, part_resp.text

    bundle_resp = client.get(f"/api/projects/{project_id}/export-bundle-json", headers=headers)
    assert bundle_resp.status_code == 200, bundle_resp.text
    payload = bundle_resp.json()
    part_summary = payload["bundle_summary"]["discrepancies"]["per_part"][0]

    assert payload["bundle_summary"]["annotations"]["total"] == 1
    assert payload["bundle_summary"]["overlays"]["configured_layers"] == 1
    assert payload["bundle_summary"]["overlays"]["segmentation_runs"] == 1
    assert payload["bundle_summary"]["measurements"]["ai_runs"] == 1
    assert part_summary["counts"]["dropped_non_object_metadata_items"] == 5
    assert "metadata_items_dropped_non_object" in part_summary["discrepancy_codes"]
    assert payload["bundle_summary"]["discrepancies"]["parts_with_discrepancies"] == 1


def test_project_bundle_json_forbidden_for_non_group_member(client):
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": "Bundle JSON Access",
            "description": "access check",
            "meta_group_id": "bundle-json-private",
            "project_type": "PT2",
        },
    )
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    with patch("routers.export.is_user_in_group", return_value=False):
        bundle_resp = client.get(f"/api/projects/{project_id}/export-bundle-json")

    assert bundle_resp.status_code == 403


def test_project_bundle_archive_supports_progressive_users_per_project_type(client):
    project_types = ("PT1", "PT2", "PT3")
    scenarios = (
        {
            "user": "basic",
            "level": 1,
            "part_count": 1,
            "image_count": 1,
        },
        {
            "user": "intermediate",
            "level": 2,
            "part_count": 2,
            "image_count": 2,
        },
        {
            "user": "advanced",
            "level": 3,
            "part_count": 3,
            "image_count": 3,
        },
    )

    for project_type in project_types:
        for scenario in scenarios:
            group = f"bundle-archive-{project_type}-{scenario['user']}"
            headers = {"X-Forwarded-Email": f"{scenario['user']}@{group}.test"}

            project_resp = client.post(
                "/api/projects/",
                json={
                    "name": f"Bundle Archive {project_type} {scenario['user']}",
                    "description": "bundle archive coverage",
                    "meta_group_id": group,
                    "project_type": project_type,
                },
                headers=headers,
            )
            assert project_resp.status_code == 201, project_resp.text
            project_id = project_resp.json()["id"]

            for idx in range(scenario["part_count"]):
                batch_resp = client.post(
                    f"/api/projects/{project_id}/batches",
                    json={"name": f"batch-{idx}", "description": f"batch {idx}"},
                    headers=headers,
                )
                assert batch_resp.status_code == 201, batch_resp.text
                batch_id = batch_resp.json()["id"]

                part_resp = client.post(
                    f"/api/projects/{project_id}/parts",
                    json={
                        "serial_number": f"{project_type}-{scenario['level']}-{idx}",
                        "display_name": f"part-{idx}",
                        "batch_id": batch_id,
                        "metadata": {
                            "synthetic_level": scenario["level"],
                            "annotations": [{"id": f"ann-{idx}", "defect_class": "scratch", "modality": "rgb"}],
                            "overlay_layers": [{"id": f"overlay-{idx}", "label": "Mask", "color": "#22c55e"}],
                            "measurement_runs": [{"run_id": f"measure-{idx}", "status": "completed"}],
                        },
                    },
                    headers=headers,
                )
                assert part_resp.status_code == 201, part_resp.text

            for image_idx in range(scenario["image_count"]):
                files = {
                    "file": (
                        f"{project_type}_{scenario['user']}_{image_idx}.png",
                        b"synthetic-image-data",
                        "image/png",
                    )
                }
                image_resp = client.post(
                    f"/api/projects/{project_id}/images",
                    files=files,
                    data={"metadata": _json.dumps({"slot": image_idx, "scenario": scenario["user"]})},
                    headers=headers,
                )
                assert image_resp.status_code == 201, image_resp.text

            bundle_resp = client.get(f"/api/projects/{project_id}/export-bundle", headers=headers)
            assert bundle_resp.status_code == 200, bundle_resp.text
            assert bundle_resp.headers["content-type"].startswith("application/zip")
            assert ".zip" in bundle_resp.headers.get("content-disposition", "")

            with zipfile.ZipFile(io.BytesIO(bundle_resp.content)) as archive:
                names = archive.namelist()
                assert "export-manifest.json" in names
                assert "export-manifest.toml" in names
                assert "project-configuration.toml" in names
                assert "project-metadata.toml" in names
                assert "images.toml" in names
                assert "parts.toml" in names
                assert "created-overlays.toml" in names
                manifest = _json.loads(archive.read("export-manifest.json").decode("utf-8"))
                manifest_toml = archive.read("export-manifest.toml").decode("utf-8")
                parts_toml = archive.read("parts.toml").decode("utf-8")

            assert manifest["project"]["project_type"] == project_type
            assert manifest["export"]["options"]["include_project_configuration"] is True
            assert "[project]" in manifest_toml
            assert "[[image_references]]" in manifest_toml
            assert "[[parts]]" in parts_toml
            assert manifest["bundle_summary"]["parts"]["total"] == scenario["part_count"]
            assert manifest["bundle_summary"]["images"]["total"] == scenario["image_count"]
            assert len(manifest["bundle_summary"]["overlays"]["records"]) == scenario["part_count"]
            assert len(manifest["bundle_summary"]["measurements"]["records"]) == scenario["part_count"]
            assert len(manifest["image_references"]) == scenario["image_count"]


def test_project_bundle_archive_respects_export_option_flags(client):
    headers = {"X-Forwarded-Email": "bundle-options@bundle-options.test"}
    project_resp = client.post(
        "/api/projects/",
        json={
            "name": "Bundle Option Flags",
            "description": "option flag coverage",
            "meta_group_id": "bundle-options",
            "project_type": "PT2",
        },
        headers=headers,
    )
    assert project_resp.status_code == 201, project_resp.text
    project_id = project_resp.json()["id"]

    image_resp = client.post(
        f"/api/projects/{project_id}/images",
        files={"file": ("option_image.png", b"image-bytes", "image/png")},
        data={"metadata": _json.dumps({"overlay": False})},
        headers=headers,
    )
    assert image_resp.status_code == 201, image_resp.text

    bundle_resp = client.get(
        (
            f"/api/projects/{project_id}/export-bundle"
            "?include_images=false"
            "&include_metadata=false"
            "&include_created_overlays=false"
            "&include_project_configuration=false"
        ),
        headers=headers,
    )
    assert bundle_resp.status_code == 200, bundle_resp.text
    with zipfile.ZipFile(io.BytesIO(bundle_resp.content)) as archive:
        names = archive.namelist()
        manifest = _json.loads(archive.read("export-manifest.json").decode("utf-8"))

    assert "export-manifest.toml" in names
    assert "project-configuration.toml" not in names
    assert "project-metadata.toml" not in names
    assert "images.toml" not in names
    assert "parts.toml" not in names
    assert "created-overlays.toml" not in names
    assert manifest["export"]["options"]["include_images"] is False
    assert manifest["image_references"][0]["archive_path"] == ""


# ---------------------------------------------------------------------------
# Unit tests for _build_workbook helper
# ---------------------------------------------------------------------------

class TestBuildWorkbook:
    """Tests for the _build_workbook helper function."""

    def _load(self, wb):
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook
        return load_workbook(buf)

    def test_empty_rows_no_meta_keys_produces_six_columns(self):
        wb = _build_workbook("Test", [], [])
        ws = self._load(wb).active
        assert ws.max_row == 1
        # Columns: Filename, Review Status, Reviewer, Review Date, Image Classes, Comment
        assert ws.cell(row=1, column=1).value == "Filename"
        assert ws.cell(row=1, column=2).value == "Review Status"
        assert ws.cell(row=1, column=3).value == "Reviewer"
        assert ws.cell(row=1, column=4).value == "Review Date"
        assert ws.cell(row=1, column=5).value == "Image Classes"
        assert ws.cell(row=1, column=6).value == "Comment"
        assert ws.max_column == 6

    def test_meta_keys_become_columns(self):
        meta_keys = ["lot_number", "serial"]
        wb = _build_workbook("Test", [], meta_keys)
        ws = self._load(wb).active
        # Columns: Filename, lot_number, serial, Review Status, Reviewer, Review Date,
        #          Image Classes, Comment
        assert ws.cell(row=1, column=1).value == "Filename"
        assert ws.cell(row=1, column=2).value == "lot_number"
        assert ws.cell(row=1, column=3).value == "serial"
        assert ws.cell(row=1, column=4).value == "Review Status"
        assert ws.cell(row=1, column=5).value == "Reviewer"
        assert ws.cell(row=1, column=6).value == "Review Date"
        assert ws.cell(row=1, column=7).value == "Image Classes"
        assert ws.cell(row=1, column=8).value == "Comment"
        assert ws.max_column == 8

    def test_freeze_panes_set(self):
        wb = _build_workbook("Test", [{"filename": "a.png"}], [])
        ws = self._load(wb).active
        assert ws.freeze_panes == "A2"

    def test_autofilter_set_with_data(self):
        rows = [{"filename": f"img{i}.png"} for i in range(3)]
        wb = _build_workbook("Test", rows, [])
        ws = self._load(wb).active
        assert ws.auto_filter.ref is not None
        assert "A1" in ws.auto_filter.ref
        # 6 columns (Filename, Review Status, Reviewer, Review Date, Image Classes, Comment),
        # 3 data rows + 1 header
        assert "F4" in ws.auto_filter.ref

    def test_autofilter_not_set_for_empty(self):
        wb = _build_workbook("Test", [], [])
        ws = self._load(wb).active
        assert ws.auto_filter.ref is None

    def test_header_styling(self):
        wb = _build_workbook("Test", [], [])
        ws = self._load(wb).active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb is not None

    def test_column_count_no_meta_keys(self):
        wb = _build_workbook("Test", [], [])
        ws = self._load(wb).active
        assert ws.max_column == 6  # Filename + Review Status + Reviewer + Review Date + Image Classes + Comment

    def test_column_count_with_meta_keys(self):
        meta_keys = ["a", "b", "c"]
        wb = _build_workbook("Test", [], meta_keys)
        ws = self._load(wb).active
        assert ws.max_column == 9  # Filename + 3 keys + Review Status + Reviewer + Review Date + Image Classes + Comment

    def test_sheet_title(self):
        wb = _build_workbook("My Project", [], [])
        ws = self._load(wb).active
        assert ws.title == "Image Data"

    def test_multiple_rows_written(self):
        meta_keys = ["lot_number"]
        rows = [
            {"filename": "img1.png", "lot_number": "L1"},
            {"filename": "img2.png", "lot_number": "L2"},
            {"filename": "img3.png", "lot_number": "L3"},
        ]
        wb = _build_workbook("Test", rows, meta_keys)
        ws = self._load(wb).active
        assert ws.max_row == 4
        assert ws.cell(row=2, column=1).value == "img1.png"
        assert ws.cell(row=3, column=1).value == "img2.png"
        assert ws.cell(row=4, column=1).value == "img3.png"

    def test_missing_keys_default_to_empty(self):
        """Row dicts missing some keys should produce empty cells."""
        meta_keys = ["lot_number", "serial"]
        rows = [{"filename": "img.png", "lot_number": "L1"}]  # missing "serial"
        wb = _build_workbook("Test", rows, meta_keys)
        ws = self._load(wb).active
        # serial (col 3) should be empty
        assert ws.cell(row=2, column=3).value in (None, "")

    def test_metadata_values_written_correctly(self):
        meta_keys = ["status", "part"]
        rows = [{"filename": "x.png", "status": "Pass", "part": "P-001"}]
        wb = _build_workbook("Test", rows, meta_keys)
        ws = self._load(wb).active
        assert ws.cell(row=2, column=1).value == "x.png"
        assert ws.cell(row=2, column=2).value == "Pass"
        assert ws.cell(row=2, column=3).value == "P-001"

    def test_formula_injection_cells_have_quote_prefix(self):
        """Cells whose values start with formula characters must have quotePrefix set."""
        meta_keys = ["formula_field"]
        rows = [
            {"filename": "=CMD", "formula_field": "=SUM(A1)"},
            {"filename": "+safe", "formula_field": "-value"},
            {"filename": "@user", "formula_field": "normal"},
        ]
        wb = _build_workbook("Test", rows, meta_keys)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook
        ws = load_workbook(buf).active
        # Values starting with formula chars are stored as text via quotePrefix
        assert ws.cell(row=2, column=1).quotePrefix is True   # "=CMD"
        assert ws.cell(row=2, column=2).quotePrefix is True   # "=SUM(A1)"
        assert ws.cell(row=3, column=1).quotePrefix is True   # "+safe"
        assert ws.cell(row=3, column=2).quotePrefix is True   # "-value"
        assert ws.cell(row=4, column=1).quotePrefix is True   # "@user"
        # Normal values must not have quotePrefix set
        assert ws.cell(row=4, column=2).quotePrefix is False  # "normal"
        # Values are preserved verbatim
        assert ws.cell(row=2, column=1).value == "=CMD"
        assert ws.cell(row=2, column=2).value == "=SUM(A1)"

    def test_formula_injection_header_cells_have_quote_prefix(self):
        """Header cells from metadata keys starting with formula chars must have quotePrefix."""
        meta_keys = ["=evil_key", "+tricky", "@mention", "normal_key"]
        wb = _build_workbook("Test", [], meta_keys)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from openpyxl import load_workbook
        ws = load_workbook(buf).active
        # Metadata key headers are columns 2-5 (column 1 is Filename)
        assert ws.cell(row=1, column=2).value == "=evil_key"
        assert ws.cell(row=1, column=2).quotePrefix is True
        assert ws.cell(row=1, column=3).value == "+tricky"
        assert ws.cell(row=1, column=3).quotePrefix is True
        assert ws.cell(row=1, column=4).value == "@mention"
        assert ws.cell(row=1, column=4).quotePrefix is True
        # Normal header should not have quotePrefix
        assert ws.cell(row=1, column=5).value == "normal_key"
        assert ws.cell(row=1, column=5).quotePrefix is False
        # Fixed headers (Filename, etc.) should not have quotePrefix
        assert ws.cell(row=1, column=1).quotePrefix is False


def test_project_backup_import_preview_and_restore_as_new(client):
    project_resp = client.post("/api/projects/", json={
        "name": "Restorable Project",
        "description": "round trip backup",
        "meta_group_id": "test-group",
        "project_type": "PT2",
    })
    assert project_resp.status_code == 201, project_resp.text
    project = project_resp.json()
    image_resp = client.post(
        f"/api/projects/{project['id']}/images",
        files={"file": ("restore.png", b"restore-bytes", "image/png")},
        data={"metadata": _json.dumps({"serial": "RESTORE-001"})},
    )
    assert image_resp.status_code == 201, image_resp.text

    export_resp = client.get(f"/api/projects/{project['id']}/export-bundle?include_images=false")
    assert export_resp.status_code == 200, export_resp.text
    assert int(export_resp.headers["x-vista-backup-estimated-bytes"]) > 0
    with zipfile.ZipFile(io.BytesIO(export_resp.content)) as archive:
        names = archive.namelist()
        assert "manifest.json" in names
        assert f"projects/{project['id']}/project-backup.json" in names

    preview_resp = client.post(
        "/api/projects/import/preview",
        files={"file": ("project.vistabundle", export_resp.content, "application/zip")},
    )
    assert preview_resp.status_code == 200, preview_resp.text
    preview = preview_resp.json()
    assert preview["valid"] is True
    assert preview["project_count"] == 1
    assert preview["projects"][0]["images"] == 1

    import_resp = client.post(
        "/api/projects/import",
        files={"file": ("project.vistabundle", export_resp.content, "application/zip")},
        data={"mode": "restore_as_new", "confirmation": "IMPORT"},
    )
    assert import_resp.status_code == 200, import_resp.text
    payload = import_resp.json()
    assert payload["ok"] is True
    new_project_id = payload["projects_created"][0]["new_project_id"]

    restored_project_resp = client.get(f"/api/projects/{new_project_id}")
    assert restored_project_resp.status_code == 200, restored_project_resp.text
    restored_project = restored_project_resp.json()
    assert restored_project["name"].startswith("Restorable Project (Imported)")
    assert restored_project["project_type"] == "PT2"

    restored_images_resp = client.get(f"/api/projects/{new_project_id}/images")
    assert restored_images_resp.status_code == 200, restored_images_resp.text
    restored_images = restored_images_resp.json()
    assert len(restored_images) == 1
    assert restored_images[0]["filename"] == "restore.png"
    assert restored_images[0]["metadata"]["serial"] == "RESTORE-001"
    assert restored_images[0]["metadata"]["source_backup"]["image_id"] == image_resp.json()["id"]


def test_dashboard_backup_export_and_import_preview(client):
    project_resp = client.post("/api/projects/", json={
        "name": "Dashboard Backup Project",
        "description": "dashboard backup",
        "meta_group_id": "test-group",
        "project_type": "PT1",
    })
    assert project_resp.status_code == 201, project_resp.text
    project_id = project_resp.json()["id"]

    export_resp = client.post(
        "/api/dashboard/export",
        json={
            "include_images": False,
            "include_overlays": False,
            "include_ui_state": True,
            "dashboard_state": {"gallery_state": {"gallery_state_demo": {"sortBy": "name"}}},
        },
    )
    assert export_resp.status_code == 200, export_resp.text
    assert export_resp.headers["content-type"].startswith("application/vnd.vista.dashboard-backup+zip")
    assert int(export_resp.headers["x-vista-backup-estimated-bytes"]) > 0
    with zipfile.ZipFile(io.BytesIO(export_resp.content)) as archive:
        manifest = _json.loads(archive.read("manifest.json").decode("utf-8"))
        dashboard_state = _json.loads(archive.read("dashboard-state.json").decode("utf-8"))
        names = archive.namelist()
    assert manifest["format"] == "vista-dashboard-backup"
    assert manifest["project_count"] == 1
    assert dashboard_state["gallery_state"]["gallery_state_demo"]["sortBy"] == "name"
    assert f"projects/{project_id}/project-backup.json" in names

    preview_resp = client.post(
        "/api/dashboard/import/preview",
        files={"file": ("dashboard.vistabundle", export_resp.content, "application/zip")},
    )
    assert preview_resp.status_code == 200, preview_resp.text
    assert preview_resp.json()["format"] == "vista-dashboard-backup"
    assert preview_resp.json()["project_count"] == 1


def test_project_import_into_active_project_append_tags_duplicate_part_serials(client):
    group = "active-import-append"
    headers = {"X-Forwarded-Email": f"user@{group}.test"}

    source_resp = client.post("/api/projects/", json={"name": "Source", "meta_group_id": group}, headers=headers)
    target_resp = client.post("/api/projects/", json={"name": "Target", "meta_group_id": group}, headers=headers)
    assert source_resp.status_code == 201, source_resp.text
    assert target_resp.status_code == 201, target_resp.text
    source_id = source_resp.json()["id"]
    target_id = target_resp.json()["id"]

    for project_id in (source_id, target_id):
        part_resp = client.post(
            f"/api/projects/{project_id}/parts",
            json={"serial_number": "PART-001", "display_name": "Part 001"},
            headers=headers,
        )
        assert part_resp.status_code == 201, part_resp.text

    bundle_resp = client.get(f"/api/projects/{source_id}/export-bundle", headers=headers)
    assert bundle_resp.status_code == 200, bundle_resp.text

    import_resp = client.post(
        f"/api/projects/{target_id}/import",
        files={"file": ("source.zip", bundle_resp.content, "application/zip")},
        data={"mode": "append_active", "confirmation": "IMPORT"},
        headers=headers,
    )
    assert import_resp.status_code == 200, import_resp.text
    assert import_resp.json()["mode"] == "append_active"

    parts_resp = client.get(f"/api/projects/{target_id}/parts", headers=headers)
    assert parts_resp.status_code == 200, parts_resp.text
    serials = sorted(part["serial_number"] for part in parts_resp.json())
    assert serials == ["PART-001", "PART-001 (duplicate)"]


def test_project_import_into_active_project_overwrite_replaces_current_parts(client):
    group = "active-import-overwrite"
    headers = {"X-Forwarded-Email": f"user@{group}.test"}

    source_resp = client.post("/api/projects/", json={"name": "Source", "meta_group_id": group}, headers=headers)
    target_resp = client.post("/api/projects/", json={"name": "Target", "meta_group_id": group}, headers=headers)
    assert source_resp.status_code == 201, source_resp.text
    assert target_resp.status_code == 201, target_resp.text
    source_id = source_resp.json()["id"]
    target_id = target_resp.json()["id"]

    source_part_resp = client.post(
        f"/api/projects/{source_id}/parts",
        json={"serial_number": "SOURCE-001", "display_name": "Source Part"},
        headers=headers,
    )
    target_part_resp = client.post(
        f"/api/projects/{target_id}/parts",
        json={"serial_number": "TARGET-001", "display_name": "Target Part"},
        headers=headers,
    )
    assert source_part_resp.status_code == 201, source_part_resp.text
    assert target_part_resp.status_code == 201, target_part_resp.text

    bundle_resp = client.get(f"/api/projects/{source_id}/export-bundle", headers=headers)
    assert bundle_resp.status_code == 200, bundle_resp.text

    import_resp = client.post(
        f"/api/projects/{target_id}/import",
        files={"file": ("source.zip", bundle_resp.content, "application/zip")},
        data={"mode": "overwrite_active", "confirmation": "IMPORT"},
        headers=headers,
    )
    assert import_resp.status_code == 200, import_resp.text
    assert import_resp.json()["mode"] == "overwrite_active"

    parts_resp = client.get(f"/api/projects/{target_id}/parts", headers=headers)
    assert parts_resp.status_code == 200, parts_resp.text
    serials = sorted(part["serial_number"] for part in parts_resp.json())
    assert serials == ["SOURCE-001"]
