import uuid
import asyncio
import io
import json
import logging
import re
import textwrap
import zipfile
import hashlib
from collections import defaultdict
from decimal import Decimal
from datetime import datetime, timezone

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, Field, field_validator
from urllib.parse import quote, urlparse, unquote
from fastapi.responses import JSONResponse, StreamingResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func as _func
from sqlalchemy.exc import IntegrityError

from core import models, schemas
from core.config import settings
from core.database import get_db
from core.group_auth_helper import is_user_in_group
from utils.boto3_client import boto3_client
from utils.dependencies import get_accessible_projects_for_user, get_current_user
import utils.crud as crud
from utils.streaming_zip import StreamingZipEntry, iter_streaming_zip
from routers.inspection_workbench import _default_project_configuration
from services.project_report_images import build_project_report_with_images_pdf

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Export"])

PROJECT_CONFIGURATION_KEY = "inspection_workbench.project_configuration"


class S3BundleLocation(BaseModel):
    s3_url: str = Field(..., min_length=1, max_length=2048)

    @field_validator("s3_url")
    @classmethod
    def validate_s3_url(cls, value: str) -> str:
        cleaned = value.strip()
        bucket, key = _parse_backup_s3_url(cleaned)
        if not key or key.endswith("/"):
            raise ValueError("S3 URL must include a destination object key, not just a bucket or prefix.")
        return cleaned


class S3ProjectExportRequest(S3BundleLocation):
    include_images: bool = True
    include_overlays: bool = True
    include_metadata: bool = True
    include_created_overlays: bool = True
    include_project_configuration: bool = True
    include_deleted: bool = False


class S3ProjectImportRequest(S3BundleLocation):
    mode: str = "append_active"
    confirmation: str = ""


class S3DashboardExportRequest(S3BundleLocation):
    include_images: bool = True
    include_overlays: bool = True
    include_metadata: bool = True
    include_created_overlays: bool = True
    include_project_configuration: bool = True
    include_deleted: bool = False
    include_archived: bool = False
    include_ui_state: bool = True
    dashboard_state: dict = Field(default_factory=dict)
    limit: int = 1000


class S3DashboardImportRequest(S3BundleLocation):
    mode: str = "restore_as_new"
    confirmation: str = ""


def _parse_backup_s3_url(raw_url: str) -> tuple[str, str]:
    cleaned = (raw_url or "").strip()
    parsed = urlparse(cleaned)
    if parsed.scheme == "s3":
        return parsed.netloc, unquote(parsed.path.lstrip("/"))
    if parsed.scheme in {"http", "https"}:
        path_parts = [part for part in parsed.path.split("/") if part]
        if ".s3." in parsed.netloc or parsed.netloc.endswith(".s3.amazonaws.com"):
            return parsed.netloc.split(".")[0], unquote(parsed.path.lstrip("/"))
        if parsed.netloc.startswith("s3.") or ".amazonaws.com" in parsed.netloc or "localhost" in parsed.netloc or ":" in parsed.netloc:
            if not path_parts:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="S3 URL must include a bucket")
            return path_parts[0], unquote("/".join(path_parts[1:]))
        return parsed.netloc.split(".")[0], unquote(parsed.path.lstrip("/"))
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Use an s3://, http://, or https:// S3 URL")


async def _zip_entries_to_bytes(entries: list[StreamingZipEntry]) -> bytes:
    chunks = []
    async for chunk in iter_streaming_zip(entries):
        chunks.append(chunk)
    return b"".join(chunks)


async def _upload_backup_bytes_to_s3(s3_url: str, payload: bytes, content_type: str) -> dict:
    if not boto3_client:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="S3 client is not configured for this backend")
    bucket, key = _parse_backup_s3_url(s3_url)
    if not bucket or not key or key.endswith("/"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="S3 URL must include a bucket and object key")
    loop = asyncio.get_running_loop()
    try:
        await loop.run_in_executor(
            None,
            lambda: boto3_client.upload_fileobj(
                io.BytesIO(payload),
                bucket,
                key,
                ExtraArgs={"ContentType": content_type},
            ),
        )
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Unable to upload backup to S3: {exc}") from exc
    return {"ok": True, "s3_url": s3_url, "bucket": bucket, "key": key, "bytes": len(payload)}


async def _download_backup_from_s3(s3_url: str) -> io.BytesIO:
    if not boto3_client:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="S3 client is not configured for this backend")
    bucket, key = _parse_backup_s3_url(s3_url)
    if not bucket or not key:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="S3 URL must include a bucket and object key")
    buffer = io.BytesIO()
    loop = asyncio.get_running_loop()
    try:
        await loop.run_in_executor(None, lambda: boto3_client.download_fileobj(bucket, key, buffer))
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Unable to download backup from S3: {exc}") from exc
    buffer.seek(0)
    return buffer


def _tag_duplicate_filename(filename: str, occurrence: int) -> str:
    safe_filename = str(filename or "").strip() or "imported.bin"
    if occurrence <= 0:
        return safe_filename
    dot_index = safe_filename.rfind(".")
    suffix = " (duplicate)" if occurrence == 1 else f" (duplicate {occurrence})"
    if dot_index > 0:
        return f"{safe_filename[:dot_index]}{suffix}{safe_filename[dot_index:]}"
    return f"{safe_filename}{suffix}"


def _dedupe_import_filename(filename: str, used_counts: dict[str, int]) -> str:
    safe_filename = str(filename or "").strip() or "imported.bin"
    occurrence = used_counts.get(safe_filename, 0)
    used_counts[safe_filename] = occurrence + 1
    return _tag_duplicate_filename(safe_filename, occurrence)


async def _get_project_with_export_access(
    project_id: uuid.UUID,
    db: AsyncSession,
    current_user: schemas.User,
) -> models.Project:
    db_project = await crud.get_project(db=db, project_id=project_id)
    if db_project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )

    if not is_user_in_group(current_user.email, db_project.meta_group_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"User does not have access to project '{project_id}'.",
        )
    return db_project


@router.get("/projects/{project_id}/export-excel")
async def export_project_excel(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    """
    Export all project image data as a Microsoft Excel (.xlsx) file.

    Each image corresponds to one row. Columns are dynamic:
    - Filename (always first)
    - One column per unique metadata key found across all project images
    - Review Status, Reviewer, Review Date (most recent review for the image)
    - Image Classes (derived from classifications)
    - Comment (derived from comments)
    """
    db_project = await _get_project_with_export_access(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )

    # Fetch all non-deleted images for the project
    result = await db.execute(
        select(models.DataInstance)
        .where(models.DataInstance.project_id == project_id)
        .where(models.DataInstance.deleted_at.is_(None))
        .order_by(models.DataInstance.created_at.asc())
    )
    images = result.scalars().all()
    image_ids = [img.id for img in images]

    # Build a class_id -> class_name lookup
    project_classes = await crud.get_image_classes_for_project(db, project_id)
    class_lookup = {str(c.id): c.name for c in project_classes}

    # Bulk-fetch all classifications for project images (avoids N+1 queries)
    classifications_by_image: dict[str, list] = defaultdict(list)
    if image_ids:
        cls_result = await db.execute(
            select(models.ImageClassification)
            .where(models.ImageClassification.image_id.in_(image_ids))
        )
        for c in cls_result.scalars().all():
            classifications_by_image[str(c.image_id)].append(c)

    # Bulk-fetch all comments for project images (avoids N+1 queries)
    comments_by_image: dict[str, list] = defaultdict(list)
    if image_ids:
        cmt_result = await db.execute(
            select(models.ImageComment)
            .where(models.ImageComment.image_id.in_(image_ids))
            .order_by(models.ImageComment.created_at.asc())
        )
        for c in cmt_result.scalars().all():
            comments_by_image[str(c.image_id)].append(c)

    # Collect unique author IDs for batch user lookup
    author_ids = set()
    for comments in comments_by_image.values():
        for c in comments:
            if c.author_id:
                author_ids.add(c.author_id)
    for img in images:
        if img.uploader_id:
            author_ids.add(img.uploader_id)

    # Bulk-fetch the most recent review per image (status, reviewer_id, created_at)
    latest_review_by_image: dict[str, models.ImageReview] = {}
    if image_ids:
        latest_review_subq = (
            select(
                models.ImageReview.image_id,
                models.ImageReview.status,
                models.ImageReview.reviewer_id,
                models.ImageReview.created_at,
                _func.row_number().over(
                    partition_by=models.ImageReview.image_id,
                    order_by=models.ImageReview.created_at.desc(),
                ).label("rn"),
            )
            .where(models.ImageReview.image_id.in_(image_ids))
            .subquery()
        )
        rev_result = await db.execute(
            select(
                latest_review_subq.c.image_id,
                latest_review_subq.c.status,
                latest_review_subq.c.reviewer_id,
                latest_review_subq.c.created_at,
            ).where(latest_review_subq.c.rn == 1)
        )
        for row in rev_result:
            latest_review_by_image[str(row.image_id)] = row
            if row.reviewer_id:
                author_ids.add(row.reviewer_id)

    # Batch-fetch all referenced users
    user_cache: dict[str, models.User] = {}
    if author_ids:
        user_result = await db.execute(
            select(models.User)
            .where(models.User.id.in_(list(author_ids)))
        )
        for u in user_result.scalars().all():
            user_cache[str(u.id)] = u

    def get_user_display(user_id) -> str:
        if user_id is None:
            return ""
        user = user_cache.get(str(user_id))
        if user is None:
            return ""
        return user.username or user.email or ""

    # Collect all unique metadata keys across all images (in order of first appearance).
    # The "measurements" key stores internal pixel-measurement overlays and is excluded.
    _EXCLUDED_META_KEYS = {"measurements"}
    all_meta_keys: list[str] = []
    seen_keys: set[str] = set()
    for image in images:
        for key in (image.metadata_json or {}).keys():
            if key not in seen_keys and key not in _EXCLUDED_META_KEYS:
                seen_keys.add(key)
                all_meta_keys.append(key)

    # Build rows from the bulk-fetched data
    rows = []
    for image in images:
        meta = image.metadata_json or {}

        # Classifications
        class_names = []
        for c in classifications_by_image.get(str(image.id), []):
            name = class_lookup.get(str(c.class_id), "Unknown")
            class_names.append(name)

        # Comments
        comment_texts = []
        for c in comments_by_image.get(str(image.id), []):
            author = get_user_display(c.author_id)
            prefix = f"[{author}] " if author else ""
            comment_texts.append(f"{prefix}{c.text}")

        row: dict[str, str] = {"filename": image.filename or ""}
        for key in all_meta_keys:
            val = meta.get(key)
            if val is None:
                row[key] = ""
            elif isinstance(val, (dict, list)):
                row[key] = json.dumps(val)
            else:
                row[key] = str(val).strip()

        # Review fields from the most recent review
        review = latest_review_by_image.get(str(image.id))
        if review:
            row["review_status"] = review.status or ""
            row["reviewer"] = get_user_display(review.reviewer_id)
            if review.created_at:
                dt = review.created_at
                row["review_date"] = dt.strftime("%Y-%m-%d %H:%M UTC")
            else:
                row["review_date"] = ""
        else:
            row["review_status"] = ""
            row["reviewer"] = ""
            row["review_date"] = ""

        row["image_classes"] = ", ".join(class_names) if class_names else ""
        row["comment"] = " | ".join(comment_texts) if comment_texts else ""
        rows.append(row)

    # Generate Excel workbook
    wb = _build_workbook(db_project.name, rows, all_meta_keys)

    # Stream response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = "".join(
        c if c.isalnum() or c in (" ", "-", "_") else "_"
        for c in db_project.name
    ).strip()
    filename = f"{safe_name}_export.xlsx"

    logger.info(
        "Excel export generated for project",
        extra={
            "row_count": len(rows),
        },
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def _normalize_part_artifact_records(part_id, serial_number, annotations, overlay_layers, measurement_runs):
    annotation_records = []
    overlay_records = []
    measurement_records = []
    incomplete_annotations = 0
    missing_measurement_ids = 0

    for annotation in annotations:
        annotation_obj = annotation if isinstance(annotation, dict) else {}
        if not annotation_obj.get("defect_class") or not annotation_obj.get("modality"):
            incomplete_annotations += 1
        annotation_records.append(
            {
                "part_id": str(part_id),
                "part_serial_number": serial_number,
                "annotation_id": annotation_obj.get("id"),
                "defect_class": annotation_obj.get("defect_class"),
                "modality": annotation_obj.get("modality"),
                "disposition": annotation_obj.get("disposition"),
                "hidden": bool(annotation_obj.get("hidden", False)),
            }
        )

    for overlay in overlay_layers:
        overlay_obj = overlay if isinstance(overlay, dict) else {}
        overlay_records.append(
            {
                "part_id": str(part_id),
                "part_serial_number": serial_number,
                "overlay_id": overlay_obj.get("id"),
                "label": overlay_obj.get("label"),
                "color": overlay_obj.get("color"),
            }
        )

    for measurement in measurement_runs:
        measurement_obj = measurement if isinstance(measurement, dict) else {}
        if not measurement_obj.get("run_id"):
            missing_measurement_ids += 1
        measurement_records.append(
            {
                "part_id": str(part_id),
                "part_serial_number": serial_number,
                "run_id": measurement_obj.get("run_id"),
                "status": measurement_obj.get("status"),
            }
        )

    return {
        "annotation_records": annotation_records,
        "overlay_records": overlay_records,
        "measurement_records": measurement_records,
        "incomplete_annotations": incomplete_annotations,
        "missing_measurement_ids": missing_measurement_ids,
    }


def _normalize_metadata_dict_list(metadata_obj, key):
    candidate = metadata_obj.get(key)
    if not isinstance(candidate, list):
        return [], 0
    normalized = []
    dropped_count = 0
    for item in candidate:
        if isinstance(item, dict):
            normalized.append(item)
        elif key == "overlay_layers" and isinstance(item, str) and item.strip():
            normalized.append({"id": item, "label": item})
        else:
            dropped_count += 1
    return normalized, dropped_count


def _json_safe(value):
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, Decimal):
        return int(value) if value == int(value) else float(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return value


def _toml_scalar(value) -> str:
    value = _json_safe(value)
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return repr(value)
    if isinstance(value, list) and all(not isinstance(item, (dict, list)) for item in value):
        return "[" + ", ".join(_toml_scalar(item) for item in value) + "]"
    if value is None:
        return '""'
    if isinstance(value, (dict, list)):
        return json.dumps(json.dumps(value, indent=2, sort_keys=True))
    return json.dumps(str(value))


def _toml_table(name: str, values: dict) -> str:
    lines = [f"[{name}]"]
    for key in sorted(values.keys()):
        safe_key = re.sub(r"[^A-Za-z0-9_-]", "_", str(key))
        lines.append(f"{safe_key} = {_toml_scalar(values[key])}")
    return "\n".join(lines)


def _toml_array_table(name: str, rows: list[dict]) -> str:
    blocks = []
    for row in rows:
        lines = [f"[[{name}]]"]
        for key in sorted(row.keys()):
            safe_key = re.sub(r"[^A-Za-z0-9_-]", "_", str(key))
            lines.append(f"{safe_key} = {_toml_scalar(row[key])}")
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks)


def _build_toml_document(*, tables: dict[str, dict] | None = None, arrays: dict[str, list[dict]] | None = None) -> str:
    blocks = []
    for name, values in (tables or {}).items():
        blocks.append(_toml_table(name, values))
    for name, rows in (arrays or {}).items():
        blocks.append(_toml_array_table(name, rows))
    return "\n\n".join(block for block in blocks if block).strip() + "\n"


def _safe_export_name(value: str, fallback: str) -> str:
    base = str(value or fallback).split("/")[-1].split("\\")[-1].strip()
    base = re.sub(r"[^A-Za-z0-9._ -]+", "_", base)
    base = re.sub(r"\s+", " ", base).strip(" .")
    return base or fallback


def _dedupe_archive_path(directory: str, filename: str, used_paths: set[str]) -> str:
    safe_name = _safe_export_name(filename, "artifact.bin")
    stem, dot, suffix = safe_name.rpartition(".")
    if not dot:
        stem, suffix = safe_name, ""
    candidate = f"{directory}/{safe_name}"
    index = 2
    while candidate in used_paths:
        candidate_name = f"{stem}-{index}.{suffix}" if suffix else f"{stem}-{index}"
        candidate = f"{directory}/{candidate_name}"
        index += 1
    used_paths.add(candidate)
    return candidate


def _image_record_is_overlay(filename: str, metadata: dict | None, part_overlay_filenames: set[str]) -> bool:
    metadata = metadata if isinstance(metadata, dict) else {}
    if bool(metadata.get("overlay")):
        return True
    filename = str(filename or "").strip()
    if filename in part_overlay_filenames:
        return True
    image_role = str(metadata.get("artifact_type") or metadata.get("role") or "").lower()
    return "overlay" in image_role


async def _read_storage_object_bytes(object_storage_key: str) -> bytes | None:
    if not boto3_client or not object_storage_key:
        return None
    buffer = io.BytesIO()
    loop = asyncio.get_running_loop()
    try:
        await loop.run_in_executor(
            None,
            lambda: boto3_client.download_fileobj(settings.S3_BUCKET, object_storage_key, buffer),
        )
    except Exception as exc:
        logger.warning(
            "Project export could not read storage object",
            extra={"object_storage_key": object_storage_key, "error": str(exc)},
        )
        return None
    return buffer.getvalue()




def _json_bytes(payload: dict | list) -> bytes:
    return json.dumps(_json_safe(payload), indent=2, sort_keys=True).encode("utf-8")


def _read_storage_object_bytes_sync(object_storage_key: str) -> bytes | None:
    if not boto3_client or not object_storage_key:
        return None
    buffer = io.BytesIO()
    try:
        boto3_client.download_fileobj(settings.S3_BUCKET, object_storage_key, buffer)
    except Exception as exc:
        logger.warning(
            "Project backup could not read storage object",
            extra={"object_storage_key": object_storage_key, "error": str(exc)},
        )
        return None
    return buffer.getvalue()


def _write_storage_object_bytes_sync(object_storage_key: str, payload: bytes, content_type: str | None) -> bool:
    if not boto3_client or not object_storage_key:
        return False
    try:
        boto3_client.upload_fileobj(
            io.BytesIO(payload),
            settings.S3_BUCKET,
            object_storage_key,
            ExtraArgs={"ContentType": content_type or "application/octet-stream"},
        )
    except Exception as exc:
        logger.warning(
            "Project import could not write storage object",
            extra={"object_storage_key": object_storage_key, "error": str(exc)},
        )
        return False
    return True


def _artifact_content_factory(image_ref: dict, manifest_payload: dict, artifact_counts: dict, missing_artifacts: list[dict]):
    def _factory() -> bytes:
        object_storage_key = image_ref.get("object_storage_key") or ""
        file_bytes = _read_storage_object_bytes_sync(object_storage_key)
        if file_bytes is None:
            artifact_counts["files_missing"] += 1
            missing_artifacts.append({
                "image_id": image_ref.get("image_id", ""),
                "filename": image_ref.get("filename", ""),
                "archive_path": image_ref.get("archive_path", ""),
                "object_storage_key": object_storage_key,
            })
            return b""
        artifact_counts["files_written"] += 1
        checksums = manifest_payload.setdefault("checksums", {})
        checksums[image_ref.get("archive_path") or ""] = hashlib.sha256(file_bytes).hexdigest()
        return file_bytes
    return _factory


async def _collect_project_backup_payload(
    *,
    project_id: uuid.UUID,
    db: AsyncSession,
    db_project: models.Project,
    current_user: schemas.User,
    include_images: bool = True,
    include_overlays: bool = True,
    include_metadata: bool = True,
    include_created_overlays: bool = True,
    include_project_configuration: bool = True,
    include_deleted: bool = False,
) -> dict:
    """Collect a restorable project backup payload plus artifact references."""

    bundle_json_response = await export_project_bundle_json(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )
    bundle_payload = json.loads(bundle_json_response.body.decode("utf-8"))

    image_query = (
        select(models.DataInstance)
        .where(models.DataInstance.project_id == project_id)
        .order_by(models.DataInstance.created_at.asc(), models.DataInstance.id.asc())
    )
    if not include_deleted:
        image_query = image_query.where(models.DataInstance.deleted_at.is_(None))
    image_rows = (await db.execute(image_query)).scalars().all()

    group_rows = (await db.execute(
        select(models.ImageGroup)
        .where(models.ImageGroup.project_id == project_id)
        .order_by(models.ImageGroup.identifier.asc(), models.ImageGroup.id.asc())
    )).scalars().all()
    batch_rows = (await db.execute(
        select(models.InspectionBatch)
        .where(models.InspectionBatch.project_id == project_id)
        .order_by(models.InspectionBatch.name.asc(), models.InspectionBatch.id.asc())
    )).scalars().all()
    part_rows = (await db.execute(
        select(models.InspectionPart)
        .where(models.InspectionPart.project_id == project_id)
        .order_by(models.InspectionPart.serial_number.asc(), models.InspectionPart.id.asc())
    )).scalars().all()
    project_metadata_rows = (await db.execute(
        select(models.ProjectMetadata)
        .where(models.ProjectMetadata.project_id == project_id)
        .order_by(models.ProjectMetadata.key.asc(), models.ProjectMetadata.id.asc())
    )).scalars().all()
    image_class_rows = (await db.execute(
        select(models.ImageClass)
        .where(models.ImageClass.project_id == project_id)
        .order_by(models.ImageClass.name.asc(), models.ImageClass.id.asc())
    )).scalars().all()

    image_ids = [image.id for image in image_rows]
    image_class_ids = [image_class.id for image_class in image_class_rows]
    classification_rows = []
    comment_rows = []
    review_rows = []
    ml_analysis_rows = []
    ml_annotation_rows = []
    if image_ids:
        classification_rows = (await db.execute(
            select(models.ImageClassification)
            .where(models.ImageClassification.image_id.in_(image_ids))
            .order_by(models.ImageClassification.created_at.asc(), models.ImageClassification.id.asc())
        )).scalars().all()
        comment_rows = (await db.execute(
            select(models.ImageComment)
            .where(models.ImageComment.image_id.in_(image_ids))
            .order_by(models.ImageComment.created_at.asc(), models.ImageComment.id.asc())
        )).scalars().all()
        review_rows = (await db.execute(
            select(models.ImageReview)
            .where(models.ImageReview.image_id.in_(image_ids))
            .order_by(models.ImageReview.created_at.asc(), models.ImageReview.id.asc())
        )).scalars().all()
        ml_analysis_rows = (await db.execute(
            select(models.MLAnalysis)
            .where(models.MLAnalysis.image_id.in_(image_ids))
            .order_by(models.MLAnalysis.created_at.asc(), models.MLAnalysis.id.asc())
        )).scalars().all()
    ml_analysis_ids = [analysis.id for analysis in ml_analysis_rows]
    if ml_analysis_ids:
        ml_annotation_rows = (await db.execute(
            select(models.MLAnnotation)
            .where(models.MLAnnotation.analysis_id.in_(ml_analysis_ids))
            .order_by(models.MLAnnotation.created_at.asc(), models.MLAnnotation.id.asc())
        )).scalars().all()

    user_ids = set()
    for image in image_rows:
        if image.uploader_id:
            user_ids.add(image.uploader_id)
        if image.deleted_by_user_id:
            user_ids.add(image.deleted_by_user_id)
    for classification in classification_rows:
        if classification.created_by_id:
            user_ids.add(classification.created_by_id)
    for comment in comment_rows:
        if comment.author_id:
            user_ids.add(comment.author_id)
    for review in review_rows:
        if review.reviewer_id:
            user_ids.add(review.reviewer_id)
    for analysis in ml_analysis_rows:
        if analysis.requested_by_id:
            user_ids.add(analysis.requested_by_id)
    user_rows = []
    if user_ids:
        user_rows = (await db.execute(
            select(models.User).where(models.User.id.in_(list(user_ids))).order_by(models.User.email.asc())
        )).scalars().all()

    project_configuration = _default_project_configuration(db_project.project_type)
    for metadata in project_metadata_rows:
        if metadata.key == PROJECT_CONFIGURATION_KEY and isinstance(metadata.value, dict):
            project_configuration = {**project_configuration, **metadata.value}
            break

    part_overlay_filenames: set[str] = set()
    for part in part_rows:
        metadata_obj = part.metadata_json if isinstance(part.metadata_json, dict) else {}
        source_images = metadata_obj.get("source_images")
        if isinstance(source_images, list):
            for record in source_images:
                if isinstance(record, dict) and record.get("overlay") and record.get("filename"):
                    part_overlay_filenames.add(str(record["filename"]))

    used_paths: set[str] = set()
    image_refs = []
    artifact_counts = {
        "images_requested": 0,
        "overlays_requested": 0,
        "files_written": 0,
        "files_missing": 0,
    }
    for image in image_rows:
        is_overlay = _image_record_is_overlay(image.filename, image.metadata_json, part_overlay_filenames)
        should_include_file = include_overlays if is_overlay else include_images
        archive_path = ""
        if should_include_file:
            artifact_counts["overlays_requested" if is_overlay else "images_requested"] += 1
            archive_path = _dedupe_archive_path(
                f"projects/{project_id}/artifacts/overlays" if is_overlay else f"projects/{project_id}/artifacts/images",
                image.filename or f"{image.id}.bin",
                used_paths,
            )
        image_refs.append({
            "image_id": str(image.id),
            "filename": image.filename,
            "archive_path": archive_path,
            "artifact_kind": "overlay" if is_overlay else "image",
            "object_storage_key": image.object_storage_key,
            "size_bytes": image.size_bytes,
            "content_type": image.content_type or "",
            "uploaded_by": image.uploaded_by_user_id or "",
            "uploader_id": str(image.uploader_id) if image.uploader_id else "",
            "group_id": str(image.group_id) if image.group_id else "",
            "created_at": image.created_at,
            "updated_at": image.updated_at,
            "deleted_at": image.deleted_at,
            "deletion_reason": image.deletion_reason,
            "storage_deleted": image.storage_deleted,
            "metadata_json": image.metadata_json if isinstance(image.metadata_json, dict) else {},
        })

    backup_payload = {
        "project": {
            "id": str(db_project.id),
            "name": db_project.name,
            "description": db_project.description or "",
            "meta_group_id": db_project.meta_group_id,
            "project_type": db_project.project_type,
            "created_by": db_project.created_by or "",
            "is_archived": bool(db_project.is_archived),
            "archived_at": db_project.archived_at,
            "created_at": db_project.created_at,
            "updated_at": db_project.updated_at,
        },
        "users": [
            {
                "id": str(user.id),
                "email": user.email,
                "username": user.username,
                "is_active": user.is_active,
            }
            for user in user_rows
        ],
        "image_groups": [
            {
                "id": str(group.id),
                "identifier": group.identifier,
                "display_name": group.display_name or "",
                "created_at": group.created_at,
                "updated_at": group.updated_at,
            }
            for group in group_rows
        ],
        "project_metadata": [
            {
                "id": str(metadata.id),
                "key": metadata.key,
                "value_json": metadata.value,
                "created_at": metadata.created_at,
                "updated_at": metadata.updated_at,
            }
            for metadata in project_metadata_rows
        ],
        "project_configuration": project_configuration,
        "images": image_refs,
        "batches": [
            {
                "id": str(batch.id),
                "name": batch.name,
                "description": batch.description or "",
                "owner": batch.owner or "",
                "status": batch.status,
                "created_at": batch.created_at,
                "updated_at": batch.updated_at,
            }
            for batch in batch_rows
        ],
        "parts": [
            {
                "id": str(part.id),
                "batch_id": str(part.batch_id) if part.batch_id else "",
                "serial_number": part.serial_number,
                "display_name": part.display_name or "",
                "review_state": part.review_state,
                "metadata_json": part.metadata_json if isinstance(part.metadata_json, dict) else {},
                "created_at": part.created_at,
                "updated_at": part.updated_at,
            }
            for part in part_rows
        ],
        "image_classes": [
            {
                "id": str(image_class.id),
                "name": image_class.name,
                "description": image_class.description or "",
                "created_at": image_class.created_at,
                "updated_at": image_class.updated_at,
            }
            for image_class in image_class_rows
        ],
        "classifications": [
            {
                "id": str(classification.id),
                "image_id": str(classification.image_id),
                "class_id": str(classification.class_id),
                "created_by_id": str(classification.created_by_id),
                "created_at": classification.created_at,
                "updated_at": classification.updated_at,
            }
            for classification in classification_rows
            if not image_class_ids or classification.class_id in image_class_ids
        ],
        "comments": [
            {
                "id": str(comment.id),
                "image_id": str(comment.image_id),
                "author_id": str(comment.author_id),
                "text": comment.text,
                "created_at": comment.created_at,
                "updated_at": comment.updated_at,
            }
            for comment in comment_rows
        ],
        "reviews": [
            {
                "id": str(review.id),
                "image_id": str(review.image_id),
                "project_id": str(review.project_id),
                "reviewer_id": str(review.reviewer_id),
                "status": review.status,
                "notes": review.notes,
                "created_at": review.created_at,
                "updated_at": review.updated_at,
            }
            for review in review_rows
        ],
        "ml_analyses": [
            {
                "id": str(analysis.id),
                "image_id": str(analysis.image_id),
                "model_name": analysis.model_name,
                "model_version": analysis.model_version,
                "status": analysis.status,
                "error_message": analysis.error_message,
                "parameters": analysis.parameters,
                "provenance": analysis.provenance,
                "requested_by_id": str(analysis.requested_by_id),
                "external_job_id": analysis.external_job_id,
                "priority": analysis.priority,
                "created_at": analysis.created_at,
                "started_at": analysis.started_at,
                "completed_at": analysis.completed_at,
                "updated_at": analysis.updated_at,
            }
            for analysis in ml_analysis_rows
        ],
        "ml_annotations": [
            {
                "id": str(annotation.id),
                "analysis_id": str(annotation.analysis_id),
                "annotation_type": annotation.annotation_type,
                "class_name": annotation.class_name,
                "confidence": annotation.confidence,
                "data": annotation.data,
                "storage_path": annotation.storage_path,
                "ordering": annotation.ordering,
                "created_at": annotation.created_at,
            }
            for annotation in ml_annotation_rows
        ],
    }

    manifest_payload = {
        "format": "vista-project-backup",
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_by": current_user.email,
        "scope": "project",
        "project": backup_payload["project"],
        "bundle_summary": bundle_payload.get("bundle_summary", {}),
        "artifact_counts": artifact_counts,
        "missing_artifacts": [],
        "options": {
            "include_images": include_images,
            "include_overlays": include_overlays,
            "include_metadata": include_metadata,
            "include_created_overlays": include_created_overlays,
            "include_project_configuration": include_project_configuration,
            "include_deleted": include_deleted,
        },
    }

    return {
        "project_id": str(project_id),
        "backup_payload": backup_payload,
        "legacy_bundle_payload": bundle_payload,
        "manifest_payload": manifest_payload,
        "image_refs": image_refs,
        "artifact_counts": artifact_counts,
        "missing_artifacts": manifest_payload["missing_artifacts"],
    }




def _estimate_project_backup_size_bytes(context: dict) -> int:
    """Return an approximate uncompressed backup size for progress UI."""

    total = 0
    for image_ref in context.get("image_refs", []):
        if image_ref.get("archive_path"):
            total += int(image_ref.get("size_bytes") or 0)
    # Account for JSON/TOML sidecars.  This is intentionally approximate because
    # the final ZIP stream is compressed and does not have a reliable content
    # length before generation finishes.
    total += len(_json_bytes(context.get("backup_payload", {})))
    total += len(_json_bytes(context.get("manifest_payload", {})))
    if context.get("legacy_bundle_payload"):
        total += len(_json_bytes(context.get("legacy_bundle_payload", {})))
    return max(total, 1)


def _estimate_dashboard_backup_size_bytes(contexts: list[dict], dashboard_state: dict) -> int:
    return max(
        sum(_estimate_project_backup_size_bytes(context) for context in contexts)
        + len(_json_bytes(dashboard_state))
        + 2048,
        1,
    )


def _project_backup_entries(context: dict, *, include_legacy_files: bool = True, include_root_manifest: bool = True) -> list[StreamingZipEntry]:
    project_id = context["project_id"]
    backup_payload = context["backup_payload"]
    manifest_payload = context["manifest_payload"]
    artifact_counts = context["artifact_counts"]
    missing_artifacts = context["missing_artifacts"]

    entries: list[StreamingZipEntry] = []
    for image_ref in context["image_refs"]:
        archive_path = image_ref.get("archive_path")
        if archive_path:
            entries.append(StreamingZipEntry(
                archive_path,
                _artifact_content_factory(image_ref, manifest_payload, artifact_counts, missing_artifacts),
            ))

    if include_root_manifest:
        entries.append(StreamingZipEntry("manifest.json", lambda: _json_bytes(manifest_payload)))

    entries.extend([
        StreamingZipEntry(f"projects/{project_id}/project-backup.json", lambda: _json_bytes(backup_payload)),
        StreamingZipEntry(f"projects/{project_id}/project.json", lambda: _json_bytes(backup_payload["project"])),
        StreamingZipEntry(f"projects/{project_id}/images.json", lambda: _json_bytes(backup_payload["images"])),
        StreamingZipEntry(f"projects/{project_id}/project-metadata.json", lambda: _json_bytes(backup_payload["project_metadata"])),
        StreamingZipEntry(f"projects/{project_id}/dashboard-state.json", lambda: _json_bytes({})),
    ])

    if include_legacy_files:
        bundle_payload = context["legacy_bundle_payload"]
        entries.extend([
            StreamingZipEntry("export-manifest.json", lambda: _json_bytes({
                "project": manifest_payload["project"],
                "bundle_summary": manifest_payload["bundle_summary"],
                "export": {
                    "format": "vista-project-export",
                    "version": 1,
                    "generated_at": manifest_payload["generated_at"],
                    "generated_by": manifest_payload["generated_by"],
                    "options": manifest_payload["options"],
                    "artifact_counts": artifact_counts,
                    "missing_artifacts": missing_artifacts,
                },
                "image_references": backup_payload["images"],
            })),
            StreamingZipEntry("export-manifest.toml", lambda: _build_toml_document(
                tables={
                    "project": _json_safe(manifest_payload["project"]),
                    "export": _json_safe({
                        "format": "vista-project-export",
                        "version": 1,
                        "generated_at": manifest_payload["generated_at"],
                        "generated_by": manifest_payload["generated_by"],
                    }),
                    "bundle_summary": _json_safe(manifest_payload["bundle_summary"]),
                },
                arrays={"image_references": _json_safe(backup_payload["images"])},
            )),
        ])
        options = manifest_payload.get("options", {})
        if options.get("include_project_configuration", True):
            entries.append(StreamingZipEntry("project-configuration.toml", lambda: _build_toml_document(
                tables={
                    "project": _json_safe(manifest_payload["project"]),
                    "project_configuration": {
                        "metadata_key": PROJECT_CONFIGURATION_KEY,
                        "config_json": _json_safe(backup_payload["project_configuration"]),
                    },
                }
            )))
        if options.get("include_metadata", True):
            entries.extend([
                StreamingZipEntry("project-metadata.toml", lambda: _build_toml_document(
                    arrays={"project_metadata": _json_safe(backup_payload["project_metadata"])}
                )),
                StreamingZipEntry("images.toml", lambda: _build_toml_document(
                    arrays={"images": _json_safe(backup_payload["images"])}
                )),
                StreamingZipEntry("parts.toml", lambda: _build_toml_document(
                    arrays={"batches": _json_safe(backup_payload["batches"]), "parts": _json_safe(backup_payload["parts"])}
                )),
            ])
        if options.get("include_created_overlays", True):
            entries.append(StreamingZipEntry("created-overlays.toml", lambda: _build_toml_document(
                tables={"project": _json_safe(manifest_payload["project"])},
                arrays={
                    "annotations": bundle_payload.get("bundle_summary", {}).get("annotations", {}).get("records", []),
                    "overlay_layers": bundle_payload.get("bundle_summary", {}).get("overlays", {}).get("records", []),
                    "measurement_runs": bundle_payload.get("bundle_summary", {}).get("measurements", {}).get("records", []),
                },
            )))
    return entries


def _manifest_from_zip(archive: zipfile.ZipFile) -> dict:
    if "manifest.json" not in archive.namelist():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup is missing manifest.json")
    try:
        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup manifest is not valid JSON") from exc
    if manifest.get("version") != 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported backup version")
    if manifest.get("format") not in {"vista-project-backup", "vista-dashboard-backup"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported backup format")
    return manifest


def _project_backup_paths(manifest: dict, archive: zipfile.ZipFile) -> list[str]:
    names = set(archive.namelist())
    if manifest.get("format") == "vista-project-backup":
        project_id = str(manifest.get("project", {}).get("id") or "")
        path = f"projects/{project_id}/project-backup.json"
        if path in names:
            return [path]
    paths = []
    for project in manifest.get("projects", []):
        project_id = str(project.get("id") or project.get("source_project_id") or "")
        path = f"projects/{project_id}/project-backup.json"
        if path in names:
            paths.append(path)
    if not paths:
        paths = sorted(name for name in names if name.startswith("projects/") and name.endswith("/project-backup.json"))
    if not paths:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup contains no project backup payloads")
    return paths


def _load_project_backup_payload(archive: zipfile.ZipFile, path: str) -> dict:
    try:
        payload = json.loads(archive.read(path).decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid project backup payload: {path}") from exc
    if not isinstance(payload.get("project"), dict):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Project backup payload missing project: {path}")
    return payload


async def _ensure_import_user(db: AsyncSession, source_user_id: str | None, source_user_by_id: dict[str, dict], current_user: schemas.User) -> models.User:
    source_user = source_user_by_id.get(str(source_user_id or ""), {})
    email = source_user.get("email") or current_user.email
    db_user = await crud.get_user_by_email(db, email)
    if db_user:
        return db_user
    db_user = models.User(email=email, username=source_user.get("username"), is_active=source_user.get("is_active", True))
    db.add(db_user)
    await db.flush()
    return db_user


async def _unique_import_project_name(db: AsyncSession, name: str, mode: str) -> str:
    if mode != "restore_as_new":
        return name
    base = f"{name} (Imported)"
    candidate = base
    suffix = 2
    while True:
        exists_result = await db.execute(select(_func.count()).select_from(models.Project).where(models.Project.name == candidate))
        if exists_result.scalar_one() == 0:
            return candidate
        candidate = f"{base} {suffix}"
        suffix += 1


async def _clear_project_import_data(db: AsyncSession, project_id: uuid.UUID) -> None:
    image_ids_result = await db.execute(select(models.DataInstance.id).where(models.DataInstance.project_id == project_id))
    image_ids = [row[0] for row in image_ids_result.all()]
    if image_ids:
        analysis_ids_result = await db.execute(select(models.MLAnalysis.id).where(models.MLAnalysis.image_id.in_(image_ids)))
        analysis_ids = [row[0] for row in analysis_ids_result.all()]
        if analysis_ids:
            await db.execute(delete(models.MLAnnotation).where(models.MLAnnotation.analysis_id.in_(analysis_ids)))
        await db.execute(delete(models.ImageDeletionEvent).where(models.ImageDeletionEvent.image_id.in_(image_ids)))
        await db.execute(delete(models.MLAnalysis).where(models.MLAnalysis.image_id.in_(image_ids)))
        await db.execute(delete(models.ImageClassification).where(models.ImageClassification.image_id.in_(image_ids)))
        await db.execute(delete(models.ImageComment).where(models.ImageComment.image_id.in_(image_ids)))
        await db.execute(delete(models.ImageReview).where(models.ImageReview.image_id.in_(image_ids)))
        await db.execute(delete(models.DataInstance).where(models.DataInstance.id.in_(image_ids)))

    await db.execute(delete(models.ImageClass).where(models.ImageClass.project_id == project_id))
    await db.execute(delete(models.InspectionPart).where(models.InspectionPart.project_id == project_id))
    await db.execute(delete(models.InspectionBatch).where(models.InspectionBatch.project_id == project_id))
    await db.execute(delete(models.ImageGroup).where(models.ImageGroup.project_id == project_id))
    await db.execute(delete(models.ProjectMetadata).where(models.ProjectMetadata.project_id == project_id))


async def _existing_filename_counts(db: AsyncSession, project_id: uuid.UUID) -> dict[str, int]:
    result = await db.execute(
        select(models.DataInstance.filename)
        .where(models.DataInstance.project_id == project_id)
        .where(models.DataInstance.deleted_at.is_(None))
    )
    counts: dict[str, int] = {}
    for (filename,) in result.all():
        safe_filename = str(filename or "").strip()
        if safe_filename:
            counts[safe_filename] = counts.get(safe_filename, 0) + 1
    return counts


async def _existing_value_counts(db: AsyncSession, model, column_name: str, project_id: uuid.UUID) -> dict[str, int]:
    column = getattr(model, column_name)
    result = await db.execute(select(column).where(model.project_id == project_id))
    counts: dict[str, int] = {}
    for (value,) in result.all():
        safe_value = str(value or "").strip()
        if safe_value:
            counts[safe_value] = counts.get(safe_value, 0) + 1
    return counts


async def _import_project_backup_payload(
    *,
    archive: zipfile.ZipFile,
    payload: dict,
    db: AsyncSession,
    current_user: schemas.User,
    mode: str,
    target_project: models.Project | None = None,
) -> dict:
    source_project = payload["project"]
    if target_project is None:
        target_group = source_project.get("meta_group_id") or ""
        if not is_user_in_group(current_user.email, target_group):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"User '{current_user.email}' cannot import project into group '{target_group}'.",
            )
    else:
        target_group = target_project.meta_group_id
        if not is_user_in_group(current_user.email, target_group):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"User '{current_user.email}' cannot import into project '{target_project.id}'.",
            )

    source_user_by_id = {str(user.get("id")): user for user in payload.get("users", []) if user.get("id")}
    if target_project is None:
        project_name = await _unique_import_project_name(db, source_project.get("name") or "Imported Project", mode)
        db_project = models.Project(
            name=project_name,
            description=source_project.get("description") or "",
            meta_group_id=target_group,
            project_type=source_project.get("project_type") or "PT1",
            created_by=current_user.email,
            is_archived=bool(source_project.get("is_archived", False)),
            archived_at=None,
        )
        db.add(db_project)
        await db.flush()
    else:
        db_project = target_project
        if mode == "overwrite_active":
            await _clear_project_import_data(db, db_project.id)
            db_project.project_type = source_project.get("project_type") or db_project.project_type
            db_project.description = source_project.get("description") or db_project.description
            await db.flush()

    filename_counts = await _existing_filename_counts(db, db_project.id)
    group_identifier_counts = await _existing_value_counts(db, models.ImageGroup, "identifier", db_project.id)
    metadata_key_counts = await _existing_value_counts(db, models.ProjectMetadata, "key", db_project.id)
    batch_name_counts = await _existing_value_counts(db, models.InspectionBatch, "name", db_project.id)
    part_serial_counts = await _existing_value_counts(db, models.InspectionPart, "serial_number", db_project.id)

    group_map: dict[str, uuid.UUID] = {}
    for source_group in payload.get("image_groups", []):
        identifier = _dedupe_import_filename(source_group.get("identifier") or "imported", group_identifier_counts)
        db_group = models.ImageGroup(
            project_id=db_project.id,
            identifier=identifier,
            display_name=source_group.get("display_name") or None,
        )
        db.add(db_group)
        await db.flush()
        group_map[str(source_group.get("id"))] = db_group.id

    for source_metadata in payload.get("project_metadata", []):
        key = _dedupe_import_filename(source_metadata.get("key") or "imported.metadata", metadata_key_counts)
        db.add(models.ProjectMetadata(
            project_id=db_project.id,
            key=key,
            value=source_metadata.get("value_json"),
        ))

    batch_map: dict[str, uuid.UUID] = {}
    for source_batch in payload.get("batches", []):
        batch_name = _dedupe_import_filename(source_batch.get("name") or "Imported Batch", batch_name_counts)
        db_batch = models.InspectionBatch(
            project_id=db_project.id,
            name=batch_name,
            description=source_batch.get("description") or None,
            owner=source_batch.get("owner") or None,
            status=source_batch.get("status") or "not_started",
        )
        db.add(db_batch)
        await db.flush()
        batch_map[str(source_batch.get("id"))] = db_batch.id

    part_map: dict[str, uuid.UUID] = {}
    for source_part in payload.get("parts", []):
        serial_number = _dedupe_import_filename(source_part.get("serial_number") or str(uuid.uuid4()), part_serial_counts)
        db_part = models.InspectionPart(
            project_id=db_project.id,
            batch_id=batch_map.get(str(source_part.get("batch_id"))),
            serial_number=serial_number,
            display_name=source_part.get("display_name") or None,
            review_state=source_part.get("review_state") or "unreviewed",
            metadata_json=source_part.get("metadata_json") if isinstance(source_part.get("metadata_json"), dict) else {},
        )
        db.add(db_part)
        await db.flush()
        part_map[str(source_part.get("id"))] = db_part.id

    image_map: dict[str, uuid.UUID] = {}
    files_uploaded = 0
    missing_files = 0
    for source_image in payload.get("images", []):
        new_image_id = uuid.uuid4()
        source_filename = source_image.get("filename") or f"{new_image_id}.bin"
        filename = _dedupe_import_filename(source_filename, filename_counts)
        object_storage_key = f"{db_project.id}/{new_image_id}/{filename}"
        archive_path = source_image.get("archive_path") or ""
        artifact_present = bool(archive_path and archive_path in archive.namelist())
        if artifact_present:
            artifact_bytes = archive.read(archive_path)
            if boto3_client:
                if _write_storage_object_bytes_sync(object_storage_key, artifact_bytes, source_image.get("content_type")):
                    files_uploaded += 1
                else:
                    missing_files += 1
            else:
                # Test/offline mode: preserve DB state while marking that storage was not written.
                missing_files += 1
        else:
            missing_files += 1
        metadata = source_image.get("metadata_json") if isinstance(source_image.get("metadata_json"), dict) else {}
        metadata = {
            **metadata,
            "source_backup": {
                "project_id": source_project.get("id"),
                "image_id": source_image.get("image_id"),
                "artifact_present": artifact_present,
                "original_filename": source_filename,
                "duplicate_filename_tagged": filename != source_filename,
            },
        }
        uploader = await _ensure_import_user(db, source_image.get("uploader_id"), source_user_by_id, current_user)
        db_image = models.DataInstance(
            id=new_image_id,
            project_id=db_project.id,
            group_id=group_map.get(str(source_image.get("group_id"))),
            filename=filename,
            object_storage_key=object_storage_key,
            content_type=source_image.get("content_type") or None,
            size_bytes=source_image.get("size_bytes"),
            metadata_json=metadata,
            uploaded_by_user_id=source_image.get("uploaded_by") or uploader.email,
            uploader_id=uploader.id,
        )
        db.add(db_image)
        image_map[str(source_image.get("image_id"))] = new_image_id

    class_map: dict[str, uuid.UUID] = {}
    for source_class in payload.get("image_classes", []):
        db_class = models.ImageClass(
            project_id=db_project.id,
            name=source_class.get("name") or "Imported Class",
            description=source_class.get("description") or None,
        )
        db.add(db_class)
        await db.flush()
        class_map[str(source_class.get("id"))] = db_class.id

    for source_classification in payload.get("classifications", []):
        image_id = image_map.get(str(source_classification.get("image_id")))
        class_id = class_map.get(str(source_classification.get("class_id")))
        if not image_id or not class_id:
            continue
        actor = await _ensure_import_user(db, source_classification.get("created_by_id"), source_user_by_id, current_user)
        db.add(models.ImageClassification(image_id=image_id, class_id=class_id, created_by_id=actor.id))

    for source_comment in payload.get("comments", []):
        image_id = image_map.get(str(source_comment.get("image_id")))
        if not image_id:
            continue
        author = await _ensure_import_user(db, source_comment.get("author_id"), source_user_by_id, current_user)
        db.add(models.ImageComment(image_id=image_id, author_id=author.id, text=source_comment.get("text") or ""))

    for source_review in payload.get("reviews", []):
        image_id = image_map.get(str(source_review.get("image_id")))
        if not image_id:
            continue
        reviewer = await _ensure_import_user(db, source_review.get("reviewer_id"), source_user_by_id, current_user)
        db.add(models.ImageReview(
            image_id=image_id,
            project_id=db_project.id,
            reviewer_id=reviewer.id,
            status=source_review.get("status") or "pass",
            notes=source_review.get("notes"),
        ))

    analysis_map: dict[str, uuid.UUID] = {}
    for source_analysis in payload.get("ml_analyses", []):
        image_id = image_map.get(str(source_analysis.get("image_id")))
        if not image_id:
            continue
        requester = await _ensure_import_user(db, source_analysis.get("requested_by_id"), source_user_by_id, current_user)
        db_analysis = models.MLAnalysis(
            image_id=image_id,
            model_name=source_analysis.get("model_name") or "imported_model",
            model_version=source_analysis.get("model_version") or "unknown",
            status=source_analysis.get("status") or "completed",
            error_message=source_analysis.get("error_message"),
            parameters=source_analysis.get("parameters"),
            provenance=source_analysis.get("provenance"),
            requested_by_id=requester.id,
            external_job_id=None,
            priority=source_analysis.get("priority") or 0,
        )
        db.add(db_analysis)
        await db.flush()
        analysis_map[str(source_analysis.get("id"))] = db_analysis.id

    for source_annotation in payload.get("ml_annotations", []):
        analysis_id = analysis_map.get(str(source_annotation.get("analysis_id")))
        if not analysis_id:
            continue
        db.add(models.MLAnnotation(
            analysis_id=analysis_id,
            annotation_type=source_annotation.get("annotation_type") or "classification",
            class_name=source_annotation.get("class_name"),
            confidence=source_annotation.get("confidence"),
            data=source_annotation.get("data") if isinstance(source_annotation.get("data"), dict) else {},
            storage_path=source_annotation.get("storage_path"),
            ordering=source_annotation.get("ordering"),
        ))

    await db.commit()
    return {
        "source_project_id": source_project.get("id"),
        "new_project_id": str(db_project.id),
        "name": db_project.name,
        "images_created": len(image_map),
        "files_uploaded": files_uploaded,
        "missing_files": missing_files,
    }


def _normalize_part_report_status(review_state: object) -> str:
    state = str(review_state or "").strip().lower()
    if state == "pass":
        return "pass"
    if state in {"reject_pending", "reject_confirmed", "reject"}:
        return "reject"
    return "unreviewed"


async def _build_legacy_project_report_payload(
    project_id: uuid.UUID,
    db: AsyncSession,
    db_project: models.Project,
) -> dict:
    """Preserve the report payload shipped before the concise schema."""

    async def count_rows(model, *criteria):
        result = await db.execute(
            select(_func.count()).select_from(model).where(*criteria)
        )
        return result.scalar_one()

    total_images = await count_rows(
        models.DataInstance,
        models.DataInstance.project_id == project_id,
        models.DataInstance.deleted_at.is_(None),
    )
    total_parts = await count_rows(
        models.InspectionPart,
        models.InspectionPart.project_id == project_id,
    )
    total_batches = await count_rows(
        models.InspectionBatch,
        models.InspectionBatch.project_id == project_id,
    )
    reviewed_states = ("pass", "reject_pending", "reject_confirmed")
    reviewed_parts = await count_rows(
        models.InspectionPart,
        models.InspectionPart.project_id == project_id,
        models.InspectionPart.review_state.in_(reviewed_states),
    )

    part_metadata_result = await db.execute(
        select(models.InspectionPart.metadata_json).where(
            models.InspectionPart.project_id == project_id
        )
    )
    metadata_drop_counts = {
        "annotations": 0,
        "overlay_layers": 0,
        "segmentation_runs": 0,
        "measurement_runs": 0,
    }
    for (metadata,) in part_metadata_result:
        metadata_obj = metadata if isinstance(metadata, dict) else {}
        for key in metadata_drop_counts:
            _, dropped = _normalize_metadata_dict_list(metadata_obj, key)
            metadata_drop_counts[key] += dropped

    part_rows_result = await db.execute(
        select(
            models.InspectionPart.id,
            models.InspectionPart.serial_number,
            models.InspectionPart.review_state,
            models.InspectionPart.updated_at,
            models.InspectionPart.metadata_json,
            models.InspectionBatch.owner,
        )
        .outerjoin(
            models.InspectionBatch,
            models.InspectionBatch.id == models.InspectionPart.batch_id,
        )
        .where(models.InspectionPart.project_id == project_id)
        .order_by(models.InspectionPart.serial_number.asc())
    )
    part_rows = part_rows_result.all()
    image_rows_result = await db.execute(
        select(
            models.DataInstance.id,
            models.DataInstance.filename,
            models.DataInstance.metadata_json,
        )
        .where(
            models.DataInstance.project_id == project_id,
            models.DataInstance.deleted_at.is_(None),
        )
        .order_by(models.DataInstance.created_at.asc())
    )

    image_part_map: dict[str, list[dict]] = defaultdict(list)
    for image_id, filename, metadata in image_rows_result.all():
        metadata_obj = metadata if isinstance(metadata, dict) else {}
        raw_part_id = metadata_obj.get("part_id")
        if raw_part_id:
            image_part_map[str(raw_part_id)].append(
                {"image_id": str(image_id), "filename": filename or ""}
            )

    part_assignments = []
    part_review_summary = []
    image_part_mappings = []
    part_status_counts = {"pass": 0, "reject": 0, "unreviewed": 0}
    for part_id, serial_number, review_state, updated_at, metadata, batch_owner in part_rows:
        report_status = _normalize_part_report_status(review_state)
        part_status_counts[report_status] += 1
        metadata_obj = metadata if isinstance(metadata, dict) else {}
        assigned_by = (
            metadata_obj.get("review_state_assigned_by")
            or metadata_obj.get("pass_fail_assigned_by")
            or metadata_obj.get("review_assigned_by")
            or "unknown"
        )
        assigned_at = (
            metadata_obj.get("review_state_assigned_at")
            or metadata_obj.get("pass_fail_assigned_at")
            or metadata_obj.get("review_assigned_at")
        )
        if not assigned_at and updated_at:
            assigned_at = updated_at.isoformat()
        part_assignments.append(
            {
                "part_id": str(part_id),
                "part_identifier": serial_number,
                "pass_fail": review_state,
                "review_status": report_status,
                "username": assigned_by,
                "batch_owner": batch_owner or "",
                "assigned_at": assigned_at or "",
            }
        )
        part_review_summary.append(
            {
                "part_id": str(part_id),
                "part_identifier": serial_number,
                "review_status": report_status,
                "raw_review_state": review_state,
            }
        )
        for image_record in image_part_map.get(str(part_id), []):
            image_part_mappings.append(
                {
                    "part_id": str(part_id),
                    "part_identifier": serial_number,
                    "image_id": image_record["image_id"],
                    "filename": image_record["filename"],
                }
            )

    return {
        "project": {
            "id": str(db_project.id),
            "name": db_project.name,
            "project_type": db_project.project_type,
            "meta_group_id": db_project.meta_group_id,
        },
        "summary": {
            "total_images": total_images,
            "total_batches": total_batches,
            "total_parts": total_parts,
            "reviewed_parts": reviewed_parts,
            "unreviewed_parts": max(total_parts - reviewed_parts, 0),
            "part_status_counts": part_status_counts,
            "metadata_normalization": {
                "dropped_non_object_items": metadata_drop_counts,
            },
        },
        "part_assignments": part_assignments,
        "part_review_summary": part_review_summary,
        "image_part_mappings": image_part_mappings,
    }


async def _build_project_report_payload(
    project_id: uuid.UUID,
    db: AsyncSession,
    db_project: models.Project,
    schema_version: int = 3,
) -> dict:
    part_rows_result = await db.execute(
        select(
            models.InspectionPart.id,
            models.InspectionPart.serial_number,
            models.InspectionPart.review_state,
        )
        .where(models.InspectionPart.project_id == project_id)
        .order_by(
            models.InspectionPart.serial_number.asc(),
            models.InspectionPart.id.asc(),
        )
    )
    part_rows = part_rows_result.all()

    parts = []
    part_status_counts = {"pass": 0, "reject": 0, "unreviewed": 0}
    for part_id, serial_number, review_state in part_rows:
        inspection_result = _normalize_part_report_status(review_state)
        part_status_counts[inspection_result] += 1
        part = {
            "part_id": str(part_id),
            "part_identifier": serial_number,
            "inspection_result": inspection_result,
        }
        if schema_version == 2:
            part["reviewed"] = inspection_result != "unreviewed"
        parts.append(part)

    total_parts = len(parts)
    reviewed_parts = part_status_counts["pass"] + part_status_counts["reject"]

    return {
        "schema_version": schema_version,
        "project": {
            "id": str(db_project.id),
            "name": db_project.name,
            "project_type": db_project.project_type,
            "meta_group_id": db_project.meta_group_id,
        },
        "summary": {
            "total_parts": total_parts,
            "reviewed_parts": reviewed_parts,
            "unreviewed_parts": part_status_counts["unreviewed"],
            "part_status_counts": part_status_counts,
        },
        "parts": parts,
    }


def _build_simple_report_pdf(report_payload: dict) -> bytes:
    project = report_payload.get("project", {})
    summary = report_payload.get("summary", {})
    parts = report_payload.get("parts", [])
    if not isinstance(parts, list):
        parts = []

    part_width = 60
    result_width = 10
    include_reviewed = report_payload.get("schema_version") == 2
    reviewed_width = 8
    table_header = f"{'Part':<{part_width}} | {'Result':<{result_width}}"
    if include_reviewed:
        table_header += f" | {'Reviewed':<{reviewed_width}}"
    table_rule = "-" * len(table_header)
    table_lines_per_page = 40

    def _sanitize_text(value: object) -> str:
        raw = str(value if value is not None else "")
        printable = "".join(character if ord(character) >= 32 else " " for character in raw)
        escaped = []
        for character in printable:
            try:
                character.encode("latin-1")
                escaped.append(character)
            except UnicodeEncodeError:
                codepoint = ord(character)
                escaped.append(
                    f"\\u{codepoint:04x}"
                    if codepoint <= 0xFFFF
                    else f"\\U{codepoint:08x}"
                )
        return "".join(escaped)

    def _wrap_text(value: object, width: int) -> list[str]:
        sanitized = _sanitize_text(value)
        return textwrap.wrap(
            sanitized,
            width=width,
            replace_whitespace=True,
            drop_whitespace=True,
            break_long_words=True,
            break_on_hyphens=False,
        ) or [""]

    def _format_table_row(
        identifier_lines: list[str],
        inspection_result: str,
        reviewed: str | None,
    ) -> list[str]:
        formatted_rows = []
        for line_index, identifier_line in enumerate(identifier_lines):
            formatted_row = (
                f"{identifier_line:<{part_width}} | "
                f"{inspection_result if line_index == 0 else '':<{result_width}}"
            )
            if include_reviewed:
                formatted_row += f" | {reviewed if line_index == 0 else '':<{reviewed_width}}"
            formatted_rows.append(formatted_row)
        return formatted_rows

    logical_table_rows = []
    for part in parts:
        row = part if isinstance(part, dict) else {}
        identifier_lines = _wrap_text(row.get("part_identifier", ""), part_width)
        inspection_result = _sanitize_text(row.get("inspection_result", "unreviewed"))[:result_width]
        reviewed = ("yes" if row.get("reviewed") is True else "no") if include_reviewed else None
        logical_table_rows.append((identifier_lines, inspection_result, reviewed))

    paginated_rows: list[list[str]] = []
    current_page_rows: list[str] = []
    for identifier_lines, inspection_result, reviewed in logical_table_rows:
        if len(identifier_lines) <= table_lines_per_page:
            formatted_row = _format_table_row(identifier_lines, inspection_result, reviewed)
            if current_page_rows and len(current_page_rows) + len(formatted_row) > table_lines_per_page:
                paginated_rows.append(current_page_rows)
                current_page_rows = []
            current_page_rows.extend(formatted_row)
            continue

        if current_page_rows:
            paginated_rows.append(current_page_rows)
            current_page_rows = []
        for start in range(0, len(identifier_lines), table_lines_per_page):
            identifier_chunk = identifier_lines[start:start + table_lines_per_page]
            paginated_rows.append(
                _format_table_row(identifier_chunk, inspection_result, reviewed)
            )
    if current_page_rows:
        paginated_rows.append(current_page_rows)
    if not paginated_rows:
        paginated_rows = [["(no parts)"]]

    def _prefixed_lines(prefix: str, value: object) -> list[str]:
        available_width = max(1, len(table_header) - len(prefix))
        wrapped = _wrap_text(value, available_width)
        return [f"{prefix}{wrapped[0]}", *wrapped[1:]]

    page_count = len(paginated_rows)
    page_lines = []
    for page_index, rows in enumerate(paginated_rows, start=1):
        if page_index == 1:
            lines = [
                "VISTA Inspection Report",
                *_prefixed_lines("Project: ", project.get("name", "Unknown")),
                f"Project ID: {_sanitize_text(project.get('id', 'Unknown'))}",
                f"Project Type: {_sanitize_text(project.get('project_type', 'PT1'))}",
                f"Total Parts: {summary.get('total_parts', len(parts))}",
                f"Reviewed Parts: {summary.get('reviewed_parts', 0)}",
                f"Unreviewed Parts: {summary.get('unreviewed_parts', 0)}",
                "",
            ]
        else:
            lines = [
                "VISTA Inspection Report (continued)",
                *_prefixed_lines("Project: ", project.get("name", "Unknown")),
                "",
            ]
        lines.extend([table_header, table_rule, *rows, "", f"Page {page_index} of {page_count}"])
        page_lines.append(lines)

    def _encode_pdf_lines(content_lines: list[str]) -> bytes:
        safe_lines = [
            str(line).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
            for line in content_lines
        ]
        text_lines = ["BT", "/F1 9 Tf", "54 756 Td", "12 TL"]
        for index, line in enumerate(safe_lines):
            if index > 0:
                text_lines.append("T*")
            text_lines.append(f"({line}) Tj")
        text_lines.append("ET")
        return "\n".join(text_lines).encode("latin-1", errors="replace")

    content_streams = [_encode_pdf_lines(lines) for lines in page_lines]
    page_object_ids = [4 + (index * 2) for index in range(page_count)]
    kids = " ".join(f"{object_id} 0 R" for object_id in page_object_ids)
    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        f"2 0 obj << /Type /Pages /Kids [{kids}] /Count {page_count} >> endobj\n".encode("ascii"),
        b"3 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Courier >> endobj\n",
    ]
    for page_index, content_stream in enumerate(content_streams):
        page_object_id = page_object_ids[page_index]
        content_object_id = page_object_id + 1
        objects.append(
            f"{page_object_id} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] ".encode("ascii")
            + f"/Resources << /Font << /F1 3 0 R >> >> /Contents {content_object_id} 0 R >> endobj\n".encode("ascii")
        )
        objects.append(
            f"{content_object_id} 0 obj << /Length {len(content_stream)} >> stream\n".encode("ascii")
            + content_stream
            + b"\nendstream endobj\n"
        )

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode("ascii")
    )
    return bytes(pdf)


def _build_legacy_simple_report_pdf(report_payload: dict) -> bytes:
    project = report_payload.get("project", {})
    summary = report_payload.get("summary", {})
    dropped = summary.get("metadata_normalization", {}).get("dropped_non_object_items", {})
    assignments = report_payload.get("part_assignments", [])
    part_review_summary = report_payload.get("part_review_summary", [])
    image_mappings = report_payload.get("image_part_mappings", [])
    lines = [
        "VISTA Inspection Report",
        f"Project: {project.get('name', 'Unknown')}",
        f"Project ID: {project.get('id', 'Unknown')}",
        f"Project Type: {project.get('project_type', 'PT1')}",
        f"Total Images: {summary.get('total_images', 0)}",
        f"Total Batches: {summary.get('total_batches', 0)}",
        f"Total Parts: {summary.get('total_parts', 0)}",
        f"Reviewed Parts: {summary.get('reviewed_parts', 0)}",
        f"Unreviewed Parts: {summary.get('unreviewed_parts', 0)}",
        f"Part Status Counts: pass={summary.get('part_status_counts', {}).get('pass', 0)}, "
        f"reject={summary.get('part_status_counts', {}).get('reject', 0)}, "
        f"unreviewed={summary.get('part_status_counts', {}).get('unreviewed', 0)}",
        "Dropped Metadata Items:",
    ]
    for field, count in dropped.items():
        lines.append(f"- {field or 'unknown_field'}: {count}")
    lines.append("Part Status Summary:")
    for part_status in part_review_summary[:40]:
        lines.append(
            f"- {part_status.get('part_identifier', '')}: "
            f"{part_status.get('review_status', 'unreviewed')}"
        )
    lines.append("Part Pass/Fail Assignments:")
    for assignment in assignments[:20]:
        lines.append(
            f"- {assignment.get('part_identifier', '')}: {assignment.get('pass_fail', '')} by "
            f"{assignment.get('username', '')} | owner {assignment.get('batch_owner', '')} | "
            f"{assignment.get('assigned_at', '')}"
        )
    mapping_lines = [
        "VISTA Report Image-to-Part Mapping",
        f"Project: {project.get('name', 'Unknown')}",
    ]
    for mapping in image_mappings[:40]:
        mapping_lines.append(
            f"- {mapping.get('filename', '')} -> {mapping.get('part_identifier', '')}"
        )

    def encode_lines(content_lines):
        safe_lines = [
            str(line).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
            for line in content_lines
        ]
        text_lines = ["BT", "/F1 10 Tf", "54 760 Td", "13 TL"]
        for index, line in enumerate(safe_lines):
            if index:
                text_lines.append("T*")
            text_lines.append(f"({line}) Tj")
        text_lines.append("ET")
        return "\n".join(text_lines).encode("latin-1", errors="replace")

    streams = [encode_lines(lines), encode_lines(mapping_lines)]
    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        b"2 0 obj << /Type /Pages /Kids [3 0 R 6 0 R] /Count 2 >> endobj\n",
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n",
        b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n",
        b"5 0 obj << /Length " + str(len(streams[0])).encode("ascii") + b" >> stream\n"
        + streams[0] + b"\nendstream endobj\n",
        b"6 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 7 0 R >> endobj\n",
        b"7 0 obj << /Length " + str(len(streams[1])).encode("ascii") + b" >> stream\n"
        + streams[1] + b"\nendstream endobj\n",
    ]
    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_offset}\n%%EOF".encode("ascii")
    )
    return bytes(pdf)


def _report_content_disposition(project_name: object, report_stem: str = "report") -> str:
    raw_name = str(project_name or "project").strip() or "project"
    safe_name = "".join(
        "_" if character in {"/", "\\"} or ord(character) < 32 or ord(character) == 127 else character
        for character in raw_name
    ).strip(" .") or "project"
    unicode_filename = f"{safe_name}-{report_stem}.pdf"

    ascii_stem = safe_name.encode("ascii", errors="replace").decode("ascii")
    ascii_stem = re.sub(r'[^A-Za-z0-9._ -]', "_", ascii_stem).strip(" .") or "project"
    ascii_stem = ascii_stem[:180].rstrip(" .") or "project"
    ascii_filename = f"{ascii_stem}-{report_stem}.pdf"
    encoded_filename = quote(unicode_filename, safe="!#$&+-.^_`|~")
    return (
        f'attachment; filename="{ascii_filename}"; '
        f"filename*=UTF-8''{encoded_filename}"
    )


@router.get("/projects/{project_id}/report-json")
async def export_project_report_json(
    project_id: uuid.UUID,
    schema_version: int = Query(default=3),
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    if schema_version not in {2, 3}:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="schema_version must be 2 or 3",
        )
    db_project = await _get_project_with_export_access(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )

    if schema_version == 2:
        report_payload = await _build_legacy_project_report_payload(
            project_id=project_id,
            db=db,
            db_project=db_project,
        )
    else:
        report_payload = await _build_project_report_payload(
            project_id=project_id,
            db=db,
            db_project=db_project,
            schema_version=schema_version,
        )
    return JSONResponse(content=report_payload)


@router.get("/projects/{project_id}/report-pdf")
async def export_project_report_pdf(
    project_id: uuid.UUID,
    schema_version: int = Query(default=3),
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    if schema_version not in {2, 3}:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="schema_version must be 2 or 3",
        )
    db_project = await _get_project_with_export_access(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )
    if schema_version == 2:
        report_payload = await _build_legacy_project_report_payload(
            project_id=project_id,
            db=db,
            db_project=db_project,
        )
        pdf_bytes = _build_legacy_simple_report_pdf(report_payload)
    else:
        report_payload = await _build_project_report_payload(
            project_id=project_id,
            db=db,
            db_project=db_project,
            schema_version=schema_version,
        )
        pdf_bytes = _build_simple_report_pdf(report_payload)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": _report_content_disposition(db_project.name)},
    )


@router.get("/projects/{project_id}/report-with-images-pdf")
async def export_project_report_with_images_pdf(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    db_project = await _get_project_with_export_access(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )
    report = await build_project_report_with_images_pdf(
        project_id=project_id,
        db=db,
        project=db_project,
    )
    return Response(
        content=report.pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": _report_content_disposition(
                db_project.name,
                "report-with-images",
            )
        },
    )


@router.get("/projects/{project_id}/export-bundle-json")
async def export_project_bundle_json(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    db_project = await _get_project_with_export_access(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )

    image_totals_result = await db.execute(
        select(
            _func.count().label("total_images"),
            _func.coalesce(_func.sum(models.DataInstance.size_bytes), 0).label("total_image_bytes"),
        )
        .select_from(models.DataInstance)
        .where(models.DataInstance.project_id == project_id)
        .where(models.DataInstance.deleted_at.is_(None))
    )
    image_totals = image_totals_result.one()

    part_metadata_result = await db.execute(
        select(
            models.InspectionPart.id,
            models.InspectionPart.serial_number,
            models.InspectionPart.display_name,
            models.InspectionPart.metadata_json,
        )
        .where(models.InspectionPart.project_id == project_id)
    )
    part_metadata_rows = part_metadata_result.all()

    annotations_count = 0
    overlay_layer_count = 0
    segmentation_run_count = 0
    measurement_run_count = 0
    annotation_records = []
    overlay_records = []
    measurement_records = []
    part_discrepancy_summaries = []

    for part_id, serial_number, display_name, metadata in part_metadata_rows:
        metadata_obj = metadata if isinstance(metadata, dict) else {}
        annotations, dropped_annotations = _normalize_metadata_dict_list(metadata_obj, "annotations")
        overlay_layers, dropped_overlay_layers = _normalize_metadata_dict_list(metadata_obj, "overlay_layers")
        segmentation_runs, dropped_segmentation_runs = _normalize_metadata_dict_list(metadata_obj, "segmentation_runs")
        measurement_runs, dropped_measurement_runs = _normalize_metadata_dict_list(metadata_obj, "measurement_runs")

        annotations_count += len(annotations)
        overlay_layer_count += len(overlay_layers)
        segmentation_run_count += len(segmentation_runs)
        measurement_run_count += len(measurement_runs)

        normalized_records = _normalize_part_artifact_records(
            part_id=part_id,
            serial_number=serial_number,
            annotations=annotations,
            overlay_layers=overlay_layers,
            measurement_runs=measurement_runs,
        )
        annotation_records.extend(normalized_records["annotation_records"])
        overlay_records.extend(normalized_records["overlay_records"])
        measurement_records.extend(normalized_records["measurement_records"])

        incomplete_annotations = normalized_records["incomplete_annotations"]
        missing_measurement_ids = normalized_records["missing_measurement_ids"]
        dropped_metadata_items = (
            dropped_annotations
            + dropped_overlay_layers
            + dropped_segmentation_runs
            + dropped_measurement_runs
        )

        discrepancy_codes = []
        if segmentation_runs and not overlay_layers:
            discrepancy_codes.append("missing_overlay_layers")
        if incomplete_annotations:
            discrepancy_codes.append("incomplete_annotation_fields")
        if missing_measurement_ids:
            discrepancy_codes.append("measurement_run_missing_run_id")
        if dropped_metadata_items:
            discrepancy_codes.append("metadata_items_dropped_non_object")

        part_discrepancy_summaries.append(
            {
                "part_id": str(part_id),
                "serial_number": serial_number,
                "display_name": display_name,
                "counts": {
                    "annotations": len(annotations),
                    "overlay_layers": len(overlay_layers),
                    "segmentation_runs": len(segmentation_runs),
                    "measurement_runs": len(measurement_runs),
                    "incomplete_annotations": incomplete_annotations,
                    "measurement_runs_missing_run_id": missing_measurement_ids,
                    "dropped_non_object_metadata_items": dropped_metadata_items,
                },
                "discrepancy_codes": discrepancy_codes,
            }
        )

    discrepancy_total = sum(1 for summary in part_discrepancy_summaries if summary["discrepancy_codes"])
    total_image_bytes = image_totals.total_image_bytes
    if isinstance(total_image_bytes, Decimal):
        total_image_bytes = int(total_image_bytes)

    bundle_payload = {
        "project": {
            "id": str(db_project.id),
            "name": db_project.name,
            "project_type": db_project.project_type,
            "meta_group_id": db_project.meta_group_id,
        },
        "bundle_summary": {
            "images": {
                "total": image_totals.total_images,
                "total_bytes": total_image_bytes,
            },
            "parts": {
                "total": len(part_metadata_rows),
            },
            "annotations": {
                "total": annotations_count,
                "records": annotation_records,
            },
            "overlays": {
                "configured_layers": overlay_layer_count,
                "segmentation_runs": segmentation_run_count,
                "records": overlay_records,
            },
            "measurements": {
                "ai_runs": measurement_run_count,
                "records": measurement_records,
            },
            "discrepancies": {
                "parts_with_discrepancies": discrepancy_total,
                "per_part": part_discrepancy_summaries,
            },
        },
    }
    return JSONResponse(content=bundle_payload)


@router.post("/projects/{project_id}/export-bundle/s3")
async def export_project_bundle_archive_to_s3(
    project_id: uuid.UUID,
    payload: S3ProjectExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    db_project = await _get_project_with_export_access(project_id=project_id, db=db, current_user=current_user)
    context = await _collect_project_backup_payload(
        project_id=project_id,
        db=db,
        db_project=db_project,
        current_user=current_user,
        include_images=payload.include_images,
        include_overlays=payload.include_overlays,
        include_metadata=payload.include_metadata,
        include_created_overlays=payload.include_created_overlays,
        include_project_configuration=payload.include_project_configuration,
        include_deleted=payload.include_deleted,
    )
    archive_bytes = await _zip_entries_to_bytes(_project_backup_entries(context, include_legacy_files=True))
    result = await _upload_backup_bytes_to_s3(payload.s3_url, archive_bytes, "application/zip")
    result["project_id"] = str(project_id)
    return result


@router.get("/projects/{project_id}/export-bundle")
async def export_project_bundle_archive(
    project_id: uuid.UUID,
    include_images: bool = True,
    include_overlays: bool = True,
    include_metadata: bool = True,
    include_created_overlays: bool = True,
    include_project_configuration: bool = True,
    include_deleted: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    db_project = await _get_project_with_export_access(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )
    context = await _collect_project_backup_payload(
        project_id=project_id,
        db=db,
        db_project=db_project,
        current_user=current_user,
        include_images=include_images,
        include_overlays=include_overlays,
        include_metadata=include_metadata,
        include_created_overlays=include_created_overlays,
        include_project_configuration=include_project_configuration,
        include_deleted=include_deleted,
    )
    safe_name = _safe_export_name(db_project.name, "project")
    filename = f"{safe_name}_export_bundle.zip"
    return StreamingResponse(
        iter_streaming_zip(_project_backup_entries(context, include_legacy_files=True)),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-VISTA-Backup-Estimated-Bytes": str(_estimate_project_backup_size_bytes(context)),
        },
    )


@router.post("/projects/import/preview")
async def preview_project_backup_import(
    file: UploadFile = File(...),
    current_user: schemas.User = Depends(get_current_user),
):
    del current_user
    try:
        with zipfile.ZipFile(file.file) as archive:
            manifest = _manifest_from_zip(archive)
            project_paths = _project_backup_paths(manifest, archive)
            projects = []
            missing_artifacts = []
            for path in project_paths:
                payload = _load_project_backup_payload(archive, path)
                project = payload["project"]
                images = payload.get("images", [])
                project_missing = [
                    {
                        "image_id": image.get("image_id"),
                        "filename": image.get("filename"),
                        "archive_path": image.get("archive_path"),
                    }
                    for image in images
                    if image.get("archive_path") and image.get("archive_path") not in archive.namelist()
                ]
                missing_artifacts.extend(project_missing)
                projects.append({
                    "source_project_id": project.get("id"),
                    "name": project.get("name"),
                    "project_type": project.get("project_type"),
                    "meta_group_id": project.get("meta_group_id"),
                    "images": len(images),
                    "parts": len(payload.get("parts", [])),
                    "comments": len(payload.get("comments", [])),
                    "classifications": len(payload.get("classifications", [])),
                    "reviews": len(payload.get("reviews", [])),
                    "missing_artifacts": len(project_missing),
                })
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup file is not a valid ZIP archive") from exc
    finally:
        await file.close()

    return {
        "valid": True,
        "format": manifest.get("format"),
        "version": manifest.get("version"),
        "scope": manifest.get("scope"),
        "projects": projects,
        "project_count": len(projects),
        "missing_artifacts": missing_artifacts,
        "warnings": ["Some image artifacts are missing from the backup."] if missing_artifacts else [],
    }


@router.post("/projects/import")
async def import_project_backup(
    file: UploadFile = File(...),
    mode: str = Form("restore_as_new"),
    confirmation: str = Form(""),
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    if mode != "restore_as_new":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only restore_as_new import mode is supported by this endpoint")
    if confirmation != "IMPORT":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Type IMPORT to confirm import")
    try:
        with zipfile.ZipFile(file.file) as archive:
            manifest = _manifest_from_zip(archive)
            dashboard_state = {}
            if "dashboard-state.json" in archive.namelist():
                dashboard_state = json.loads(archive.read("dashboard-state.json").decode("utf-8"))
            project_paths = _project_backup_paths(manifest, archive)
            imported_projects = []
            for path in project_paths:
                payload = _load_project_backup_payload(archive, path)
                imported_projects.append(await _import_project_backup_payload(
                    archive=archive,
                    payload=payload,
                    db=db,
                    current_user=current_user,
                    mode=mode,
                ))
    except zipfile.BadZipFile as exc:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup file is not a valid ZIP archive") from exc
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Import failed due to conflicting imported records") from exc
    except HTTPException:
        await db.rollback()
        raise
    except Exception:
        await db.rollback()
        raise
    finally:
        await file.close()

    return {
        "ok": True,
        "format": manifest.get("format"),
        "projects_created": imported_projects,
        "project_count": len(imported_projects),
        "dashboard_state": dashboard_state,
    }

@router.post("/projects/{project_id}/import/s3")
async def import_project_backup_into_active_project_from_s3(
    project_id: uuid.UUID,
    payload: S3ProjectImportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    s3_file = UploadFile(filename=payload.s3_url, file=await _download_backup_from_s3(payload.s3_url))
    return await import_project_backup_into_active_project(
        project_id=project_id,
        file=s3_file,
        mode=payload.mode,
        confirmation=payload.confirmation,
        db=db,
        current_user=current_user,
    )


@router.post("/projects/{project_id}/import")
async def import_project_backup_into_active_project(
    project_id: uuid.UUID,
    file: UploadFile = File(...),
    mode: str = Form("append_active"),
    confirmation: str = Form(""),
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    if mode not in {"append_active", "overwrite_active"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Choose append_active or overwrite_active import mode")
    if confirmation != "IMPORT":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Type IMPORT to confirm import")

    db_project = await _get_project_with_export_access(
        project_id=project_id,
        db=db,
        current_user=current_user,
    )
    try:
        with zipfile.ZipFile(file.file) as archive:
            manifest = _manifest_from_zip(archive)
            project_paths = _project_backup_paths(manifest, archive)
            if len(project_paths) != 1:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Importing into the active project requires a bundle with exactly one project",
                )
            payload = _load_project_backup_payload(archive, project_paths[0])
            imported_project = await _import_project_backup_payload(
                archive=archive,
                payload=payload,
                db=db,
                current_user=current_user,
                mode=mode,
                target_project=db_project,
            )
    except zipfile.BadZipFile as exc:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Backup file is not a valid ZIP archive") from exc
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Import failed due to conflicting imported records") from exc
    except HTTPException:
        await db.rollback()
        raise
    except Exception:
        await db.rollback()
        raise
    finally:
        await file.close()

    return {
        "ok": True,
        "format": manifest.get("format"),
        "mode": mode,
        "project": imported_project,
        "project_count": 1,
    }



async def _build_dashboard_backup_entries(
    options: dict,
    db: AsyncSession,
    current_user: schemas.User,
) -> tuple[list[StreamingZipEntry], dict, int]:
    include_archived = bool(options.get("include_archived", False))
    projects = await get_accessible_projects_for_user(
        db=db,
        user=current_user,
        skip=0,
        limit=int(options.get("limit", 1000) or 1000),
        include_archived=include_archived,
    )
    contexts = []
    for project in projects:
        contexts.append(await _collect_project_backup_payload(
            project_id=project.id,
            db=db,
            db_project=project,
            current_user=current_user,
            include_images=bool(options.get("include_images", True)),
            include_overlays=bool(options.get("include_overlays", True)),
            include_metadata=bool(options.get("include_metadata", True)),
            include_created_overlays=bool(options.get("include_created_overlays", True)),
            include_project_configuration=bool(options.get("include_project_configuration", True)),
            include_deleted=bool(options.get("include_deleted", False)),
        ))

    dashboard_state = options.get("dashboard_state") if isinstance(options.get("dashboard_state"), dict) else {}
    manifest_payload = {
        "format": "vista-dashboard-backup",
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_by": current_user.email,
        "scope": "dashboard",
        "project_count": len(contexts),
        "projects": [context["manifest_payload"]["project"] for context in contexts],
        "options": {
            "include_archived": include_archived,
            "include_images": bool(options.get("include_images", True)),
            "include_overlays": bool(options.get("include_overlays", True)),
            "include_metadata": bool(options.get("include_metadata", True)),
            "include_created_overlays": bool(options.get("include_created_overlays", True)),
            "include_project_configuration": bool(options.get("include_project_configuration", True)),
            "include_deleted": bool(options.get("include_deleted", False)),
            "include_ui_state": bool(options.get("include_ui_state", True)),
        },
    }
    entries: list[StreamingZipEntry] = []
    for context in contexts:
        entries.extend(_project_backup_entries(context, include_legacy_files=False, include_root_manifest=False))
    entries.append(StreamingZipEntry("manifest.json", lambda: _json_bytes(manifest_payload)))
    entries.append(StreamingZipEntry("dashboard-state.json", lambda: _json_bytes(dashboard_state)))
    return entries, dashboard_state, _estimate_dashboard_backup_size_bytes(contexts, dashboard_state)


@router.post("/dashboard/export/s3")
async def export_dashboard_backup_to_s3(
    payload: S3DashboardExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    entries, _dashboard_state, _estimate = await _build_dashboard_backup_entries(payload.model_dump(), db, current_user)
    archive_bytes = await _zip_entries_to_bytes(entries)
    return await _upload_backup_bytes_to_s3(payload.s3_url, archive_bytes, "application/vnd.vista.dashboard-backup+zip")


@router.post("/dashboard/export")
async def export_dashboard_backup(
    options: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    entries, _dashboard_state, estimated_bytes = await _build_dashboard_backup_entries(options, db, current_user)
    filename = f"vista-dashboard-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}.vistabundle"
    return StreamingResponse(
        iter_streaming_zip(entries),
        media_type="application/vnd.vista.dashboard-backup+zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-VISTA-Backup-Estimated-Bytes": str(estimated_bytes),
        },
    )


@router.post("/dashboard/import/preview")
async def preview_dashboard_backup_import(
    file: UploadFile = File(...),
    current_user: schemas.User = Depends(get_current_user),
):
    return await preview_project_backup_import(file=file, current_user=current_user)


@router.post("/dashboard/import/s3")
async def import_dashboard_backup_from_s3(
    payload: S3DashboardImportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    s3_file = UploadFile(filename=payload.s3_url, file=await _download_backup_from_s3(payload.s3_url))
    return await import_project_backup(
        file=s3_file,
        mode=payload.mode,
        confirmation=payload.confirmation,
        db=db,
        current_user=current_user,
    )


@router.post("/dashboard/import")
async def import_dashboard_backup(
    file: UploadFile = File(...),
    mode: str = Form("restore_as_new"),
    confirmation: str = Form(""),
    db: AsyncSession = Depends(get_db),
    current_user: schemas.User = Depends(get_current_user),
):
    return await import_project_backup(
        file=file,
        mode=mode,
        confirmation=confirmation,
        db=db,
        current_user=current_user,
    )



def _build_workbook(project_name: str, rows: list[dict], meta_keys: list[str]):
    """Build an openpyxl Workbook from the collected row data.

    Columns are dynamic:
    - Filename (always first)
    - One column per unique metadata key
    - Review Status, Reviewer, Review Date
    - Image Classes
    - Comment
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Image Data"

    # Dynamic column definitions: (header_label, width)
    columns = [("Filename", 35)]
    for key in meta_keys:
        columns.append((key, 20))
    columns.append(("Review Status", 20))
    columns.append(("Reviewer", 25))
    columns.append(("Review Date", 22))
    columns.append(("Image Classes", 30))
    columns.append(("Comment", 50))

    # The dict keys used to retrieve values from each row
    row_keys = (
        ["filename"]
        + list(meta_keys)
        + ["review_status", "reviewer", "review_date", "image_classes", "comment"]
    )

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Characters that trigger formula evaluation in Excel
    _FORMULA_CHARS = frozenset("=+-@")

    # Write headers
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        if isinstance(header, str) and header and header[0] in _FORMULA_CHARS:
            cell.quotePrefix = True
        ws.column_dimensions[cell.column_letter].width = width

    # Data styling
    data_alignment = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, key in enumerate(row_keys, start=1):
            value = row_data.get(key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            # Prevent formula injection: values starting with formula characters
            # are marked as text-prefixed so Excel does not evaluate them.
            if isinstance(value, str) and value and value[0] in _FORMULA_CHARS:
                cell.quotePrefix = True
            cell.border = thin_border

            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Add autofilter
    if rows and columns:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    return wb
